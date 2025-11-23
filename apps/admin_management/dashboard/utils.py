"""
Dashboard utility functions for pagination, Excel export, and query optimization
"""

from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Fallback to CSV if openpyxl not available
    import csv

from io import BytesIO


def paginate_queryset(queryset, page=1, page_size=100, max_page_size=100):
    """
    Paginate queryset with maximum page size limit
    
    Args:
        queryset: Django queryset
        page: Page number (default: 1)
        page_size: Items per page (default: 100, max: 100)
        max_page_size: Maximum allowed page size (default: 100)
    
    Returns:
        dict with paginated data and pagination info
    """
    # Enforce maximum page size
    if page_size > max_page_size:
        page_size = max_page_size
    
    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page)
    
    return {
        'items': list(page_obj),
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        }
    }


def filter_users_queryset(queryset, search=None, role=None, status=None):
    """
    Apply filters to users queryset
    
    Args:
        queryset: AdminUser queryset
        search: Search term for username, email, full_name
        role: Filter by admin_level
        status: Filter by status
    
    Returns:
        Filtered queryset
    """
    if search:
        queryset = queryset.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search) |
            Q(full_name__icontains=search)
        )
    
    if role:
        queryset = queryset.filter(admin_level=role)
    
    if status:
        queryset = queryset.filter(status=status)
    
    return queryset


def export_to_excel(data, headers, filename='export.xlsx', sheet_name='Data'):
    """
    Export data to Excel file (or CSV if openpyxl not available)
    
    Args:
        data: List of dictionaries or list of lists
        headers: List of header names
        filename: Output filename
        sheet_name: Sheet name
    
    Returns:
        HttpResponse with Excel/CSV file
    """
    if not OPENPYXL_AVAILABLE:
        # Fallback to CSV export
        return export_to_csv(data, headers, filename.replace('.xlsx', '.csv'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header row styling
    header_fill = PatternFill(start_color="3a6ea5ff", end_color="3a6ea5ff", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        if isinstance(row_data, dict):
            # Convert dict to list based on headers order
            row_values = [row_data.get(header, '') for header in headers]
        else:
            row_values = row_data
        
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.read())
    
    return response


def export_to_csv(data, headers, filename='export.csv'):
    """
    Export data to CSV file (fallback when openpyxl not available)
    
    Args:
        data: List of dictionaries or list of lists
        headers: List of header names
        filename: Output filename
    
    Returns:
        HttpResponse with CSV file
    """
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for row_data in data:
        if isinstance(row_data, dict):
            row_values = [row_data.get(header, '') for header in headers]
        else:
            row_values = row_data
        writer.writerow(row_values)
    
    return response


def format_relative_time(timestamp):
    """
    Format timestamp to relative time (e.g., "2 hours ago")
    
    Args:
        timestamp: datetime object
    
    Returns:
        Formatted string
    """
    if not timestamp:
        return "Never"
    
    now = timezone.now()
    diff = now - timestamp
    
    if diff.days > 0:
        return f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
    elif diff.seconds >= 3600:
        hours = diff.seconds // 3600
        return f"{hours} hour{'s' if hours > 1 else ''} ago"
    elif diff.seconds >= 60:
        minutes = diff.seconds // 60
        return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
    else:
        return "Just now"

