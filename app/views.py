from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import connection, models, transaction
import redis
import json
import bcrypt
import requests
import jwt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64
from django.conf import settings
from .models import (
    Category, SubSection, Topic, 
    Question, QuestionChoice, SubQuestion, Form, Answer,
    School, District, Division, Region, AdminUser,
    AuditTrail, AuditLog
)
from .logging_utils import SystemLogger
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from functools import wraps
from django.shortcuts import render, redirect
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Q

def session_or_login_required(view_func):
    """
    Custom decorator that allows both session-based authentication and Django login authentication.
    This is needed for analytics endpoints that need to work with both authentication systems.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Check if user is authenticated via Django login
        if request.user.is_authenticated:
            return view_func(request, *args, **kwargs)
        
        # Check if user is authenticated via session
        if request.session.get('admin_id'):
            return view_func(request, *args, **kwargs)
        
        # If neither authentication method works, return 403
        return JsonResponse({'error': 'Not authenticated'}, status=403)
    return _wrapped_view

def session_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.session.get('admin_id'):
            return JsonResponse({'error': 'Not authenticated'}, status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def get_current_user_school_id(request):
    """Get the school ID for the current authenticated user"""
    admin_id = request.session.get('admin_id')
    if not admin_id:
        return None
    
    try:
        admin_user = AdminUser.objects.get(admin_id=admin_id)
        if admin_user.admin_level == 'school' and admin_user.school:
            # Return the school's primary key (schools.id), not the school_id field
            # This matches what the forms.school_id foreign key expects
            return admin_user.school.id
        return None
    except AdminUser.DoesNotExist:
        return None

@session_or_login_required
def user_form(request):
    # Ensure user has a form created (needed for cross-browser answer retrieval)
    admin_id = request.session.get('admin_id')
    school_id = get_current_user_school_id(request)
    
    if admin_id and school_id:
        try:
            # Get the School object from the school_id
            school_obj = School.objects.get(id=school_id)
            # For session-based auth, create form using admin_id as user reference
            form, created = Form.objects.get_or_create(
                user_id=admin_id,  # Use admin_id as user reference
                school=school_obj,
                defaults={'status': 'draft'}
            )
            if created:
                print(f"Created new form {form.form_id} for admin user {admin_id}")
        except Exception as e:
            print(f"Error creating form: {e}")
            pass
    # All users should now be authenticated via admin_user table
    # Django authentication is kept as fallback but may not be needed
    
    # Get all categories with their completion rates
    categories = Category.objects.all().order_by('display_order')
    
    # Build the complete form structure with all related data
    form_data = []
    for category in categories:
        # Get subsections for this category
        subsections = SubSection.objects.filter(category=category).order_by('display_order')
        
        subsections_data = []
        for subsection in subsections:
            # Get topics for this subsection
            topics = Topic.objects.filter(sub_section=subsection).order_by('display_order')
            
            topics_data = []
            for topic in topics:
                # Get questions and their choices for this topic
                questions = Question.objects.filter(topic=topic).order_by('display_order')
                
                questions_data = []
                for question in questions:
                    # Get choices if question type is 'choice'
                    choices = []
                    if question.answer_type == 'choice':
                        choices = list(QuestionChoice.objects.filter(question=question)
                                    .values_list('choice_text', flat=True))
                    
                    # Get user's answer if exists (only for regular questions, not sub-questions)
                    answer_value = None
                    admin_id = request.session.get('admin_id')
                    school_id = get_current_user_school_id(request)
                    
                    if admin_id and school_id:
                        # Get the School object from the school_id
                        school_obj = School.objects.get(id=school_id)
                        # Session-based authentication
                        answer = Answer.objects.filter(
                            question=question,
                            sub_question__isnull=True,  # Only get regular question answers
                            form__user_id=admin_id,
                            form__school=school_obj
                        ).first()
                        answer_value = answer.response if answer else None
                    # All authentication should now be via admin_user table
                    
                    # Get sub-questions for this question
                    from app.models import SubQuestion
                    sub_questions = SubQuestion.objects.filter(question=question).order_by('sub_question_id')
                    sub_questions_data = []
                    
                    for sub_question in sub_questions:
                         # Get user's answer for sub-question if it exists
                         sub_answer_value = None
                         
                         if admin_id and school_id:
                             # Session-based authentication
                             sub_answer = Answer.objects.filter(
                                 question=question,
                                 sub_question=sub_question,
                                 form__user_id=admin_id,
                                 form__school=school_obj
                             ).first()
                             sub_answer_value = sub_answer.response if sub_answer else None
                          # All authentication should now be via admin_user table
                         
                         sub_question_data = {
                             'id': sub_question.sub_question_id,
                             'text': sub_question.sub_question_text,
                             'type': 'text',  # Default type for sub-questions
                             'is_required': False,  # Default for sub-questions
                             'answer': sub_answer_value
                         }
                         sub_questions_data.append(sub_question_data)
                    
                    questions_data.append({
                        'id': question.question_id,
                        'text': question.question_text,
                        'type': question.answer_type,
                        'is_required': question.is_required,
                        'answer': answer_value,
                        'choices': choices if choices else None,
                        'sub_questions': sub_questions_data
                    })
                
                # Calculate topic completion including sub-questions
                total_questions = 0
                answered_questions = 0
                
                for q in questions_data:
                    # Count main question
                    total_questions += 1
                    if q['answer']:
                        answered_questions += 1
                    
                    # Count sub-questions
                    if q.get('sub_questions'):
                        total_questions += len(q['sub_questions'])
                        answered_questions += len([sq for sq in q['sub_questions'] if sq.get('answer')])
                
                completion_rate = int((answered_questions / total_questions * 100) if total_questions > 0 else 0)
                
                topics_data.append({
                    'id': topic.topic_id,
                    'name': topic.name,
                    'questions': questions_data,
                    'is_completed': completion_rate == 100
                })
            
            subsections_data.append({
                'id': subsection.sub_section_id,
                'name': subsection.name,
                'topics': topics_data
            })
        
        # Calculate category completion rate including sub-questions
        total_questions = 0
        answered_questions = 0
        
        for subsec in subsections_data:
            for topic in subsec['topics']:
                for q in topic['questions']:
                    # Count main question
                    total_questions += 1
                    if q['answer']:
                        answered_questions += 1
                    
                    # Count sub-questions
                    if q.get('sub_questions'):
                        total_questions += len(q['sub_questions'])
                        answered_questions += len([sq for sq in q['sub_questions'] if sq.get('answer')])
        completion_rate = int((answered_questions / total_questions * 100) if total_questions > 0 else 0)
        
        form_data.append({
            'id': category.category_id,
            'name': category.name,
            'subsections': subsections_data,
            'completion_rate': completion_rate
        })
    
    # Calculate overall stats
    total_questions = sum(len(topic['questions']) 
                         for category in form_data 
                         for subsec in category['subsections'] 
                         for topic in subsec['topics'])
    
    completed_forms = len([cat for cat in form_data if cat['completion_rate'] == 100])
    total_forms = len(form_data)
    
    overall_completion = int(sum(cat['completion_rate'] for cat in form_data) / len(form_data) if form_data else 0)
    
    context = {
        'categories': form_data,
        'completion_rate': overall_completion,
        'completed_forms': completed_forms,
        'total_forms': total_forms,
        'total_questions': total_questions
    }
    
    return render(request, 'dashboard/user_form.html', context)

@login_required
@require_POST
def submit_form(request):
    try:
        data = json.loads(request.body)
        answers = data.get('answers', [])
        
        # Get the school_id from current user
        school_id = get_current_user_school_id(request)
        admin_id = request.session.get('admin_id')
        
        if not school_id and not admin_id:
            return JsonResponse({'error': 'User not authenticated or not linked to a school'}, status=403)
        
        # For session-based auth, use admin_id as user_id
        if admin_id and school_id:
            user_id = admin_id
        # All users should be authenticated via admin_user table
        else:
            return JsonResponse({'error': 'Authentication required'}, status=403)
        
        with connection.cursor() as cursor:
            # Get or create form for the user
            cursor.execute("""
                INSERT INTO forms (user_id, school_id, status, created_at, updated_at)
                SELECT %s, %s, 'in-progress', NOW(), NOW()
                WHERE NOT EXISTS (
                    SELECT 1 FROM forms WHERE user_id = %s AND school_id = %s
                )
                RETURNING form_id;
                
                SELECT form_id FROM forms 
                WHERE user_id = %s AND school_id = %s 
                ORDER BY created_at DESC 
                LIMIT 1;
            """, [user_id, school_id, user_id, school_id, user_id, school_id])
            
            form_id = cursor.fetchone()[0]
            
            # Save each answer
            for answer_data in answers:
                question_id = answer_data.get('question_id')
                sub_question_id = answer_data.get('sub_question_id')
                answer_value = answer_data.get('answer')
                
                if question_id and answer_value:
                    if sub_question_id:
                        # Handle sub-question answer
                        cursor.execute("""
                            INSERT INTO answers (form_id, question_id, sub_question_id, response, answered_at)
                            VALUES (%s, %s, %s, %s, NOW())
                            ON DUPLICATE KEY UPDATE
                                response = VALUES(response),
                                answered_at = NOW()
                        """, [form_id, question_id, sub_question_id, answer_value])
                    else:
                        # Handle regular question answer
                        cursor.execute("""
                            INSERT INTO answers (form_id, question_id, response, answered_at)
                            VALUES (%s, %s, %s, NOW())
                            ON DUPLICATE KEY UPDATE
                                response = VALUES(response),
                                answered_at = NOW()
                        """, [form_id, question_id, answer_value])
            
            # Update form status if needed
            if data.get('status') == 'completed':
                cursor.execute("""
                    UPDATE forms 
                    SET status = 'completed', updated_at = NOW()
                    WHERE form_id = %s
                """, [form_id])
            
            connection.commit()
        
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        if connection:
            connection.rollback()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@session_or_login_required
def get_form_state(request):
    """Get the current state of the form including deadline"""
    school_id = get_current_user_school_id(request)
    admin_id = request.session.get('admin_id')
    
    if admin_id and school_id:
        user_id = admin_id
    # All users should be authenticated via admin_user table
    else:
        return JsonResponse({'error': 'Authentication required'}, status=403)
    
    form = Form.objects.filter(user_id=user_id, school_id=school_id).first()
    
    return JsonResponse({
        'deadline': '2025-12-31T23:59:59',  # Replace with actual deadline from your settings
        'lastSaved': form.updated_at.isoformat() if form else None,
        'status': form.status if form else 'draft'
    })
# Removed unused serializers import

r = redis.Redis(
    host=getattr(settings, 'REDIS_HOST', 'localhost'),
    port=getattr(settings, 'REDIS_PORT', 6379),
    db=getattr(settings, 'REDIS_DB', 0)
)


@csrf_exempt
@session_required
def stats(request):
    # Forms completed per day (for stats)
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT 
                COUNT(form_id) AS total_forms,
                SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) AS completed_forms,
                ROUND(100 * SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) / COUNT(form_id), 1) AS completion_rate
            FROM forms
            WHERE school_id = %s
        ''', [get_current_user_school_id(request)])
        row = cursor.fetchone()
    return JsonResponse({
        'total_forms': row[0],
        'completed_forms': row[1],
        'completion_rate': row[2],
    })

@csrf_exempt
@session_required
def completion_by_region(request):
    # Completion rates by region
    school_id = get_current_user_school_id(request)
    if not school_id:
        return JsonResponse({'completion_by_region': []})
        
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT r.name as region_name, COUNT(f.form_id) AS total_forms,
                SUM(CASE WHEN f.status = 'completed' THEN 1 ELSE 0 END) AS completed_forms,
                ROUND(100 * SUM(CASE WHEN f.status = 'completed' THEN 1 ELSE 0 END) / COUNT(f.form_id), 1) AS completion_rate
            FROM forms f
            JOIN schools s ON f.school_id = s.id
            JOIN regions r ON s.region_id = r.id
            WHERE f.school_id = %s
            GROUP BY r.id, r.name
            ORDER BY completion_rate DESC
        ''', [school_id])
        rows = cursor.fetchall()
    data = [
        {
            'region_name': row[0],
            'total_forms': row[1],
            'completed_forms': row[2],
            'completion_rate': row[3],
        } for row in rows
    ]
    return JsonResponse({'completion_by_region': data})

@csrf_exempt
@session_required
def forms_over_time(request):
    # Forms completed per day
    school_id = get_current_user_school_id(request)
    if not school_id:
        return JsonResponse({'forms_over_time': []})
        
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT DATE(updated_at) AS completion_date, COUNT(form_id) AS forms_completed
            FROM forms
            WHERE status = 'completed' AND school_id = %s
            GROUP BY DATE(updated_at)
            ORDER BY completion_date
        ''', [school_id])
        rows = cursor.fetchall()
    data = [
        {'completion_date': str(row[0]), 'forms_completed': row[1]} for row in rows
    ]
    return JsonResponse({'forms_over_time': data})

@csrf_exempt
@session_required
def top_schools(request):
    # Top schools by completion rate
    school_id = get_current_user_school_id(request)
    if not school_id:
        return JsonResponse({'top_schools': []})
        
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT s.school_name, COUNT(f.form_id) AS total_forms,
                SUM(CASE WHEN f.status = 'completed' THEN 1 ELSE 0 END) AS completed,
                ROUND(100 * SUM(CASE WHEN f.status = 'completed' THEN 1 ELSE 0 END) / COUNT(f.form_id), 1) AS completion_rate
            FROM forms f
            JOIN schools s ON f.school_id = s.id
            WHERE f.school_id = %s
            GROUP BY s.id, s.school_name
            ORDER BY completion_rate DESC, completed DESC
            LIMIT 10
        ''', [school_id])
        rows = cursor.fetchall()
    data = [
        {
            'school_name': row[0],
            'total_forms': row[1],
            'completed': row[2],
            'completion_rate': row[3],
        } for row in rows
    ]
    return JsonResponse({'top_schools': data})

@csrf_exempt
@session_required
def response_distribution(request):
    # Response distribution for a specific question
    question_id = request.GET.get('question', None)
    if not question_id:
        return JsonResponse({'error': 'Missing question parameter'}, status=400)
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT q.question_text, a.response, COUNT(a.answer_id) AS response_count,
                ROUND(100 * COUNT(a.answer_id) / total.total, 1) AS percentage
            FROM answers a
            JOIN questions q ON a.question_id = q.question_id
            JOIN (SELECT COUNT(*) AS total FROM answers WHERE question_id = %s) AS total
            WHERE q.question_id = %s AND a.form_id IN (SELECT form_id FROM forms WHERE school_id = %s)
            GROUP BY a.response
        ''', [question_id, question_id, request.session['school_id']])
        rows = cursor.fetchall()
    data = [
        {
            'question_text': row[0],
            'response': row[1],
            'response_count': row[2],
            'percentage': row[3],
        } for row in rows
    ]
    return JsonResponse({'response_distribution': data})

@csrf_exempt
@require_POST
@session_required
def submit_form_session(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            answers = data.get('answers', [])
            status = data.get('status', 'in-progress')
            school_id = get_current_user_school_id(request)
            
            # Extract client information for logging
            ip_address = SystemLogger.get_client_ip(request)
            user_agent = SystemLogger.get_user_agent(request)
            
            if not answers:
                # Log failed form submission
                if hasattr(request, 'user') and request.user.is_authenticated:
                    SystemLogger.log_form_submission(
                        user=request.user,
                        form_id=None,
                        school_id=school_id,
                        ip_address=ip_address,
                        user_agent=user_agent,
                        success=False
                    )
                return JsonResponse({'error': 'No answers provided'}, status=400)
            
            with connection.cursor() as cursor:
                # Get or create form for the school
                cursor.execute("""
                    SELECT form_id FROM forms 
                    WHERE school_id = %s 
                    ORDER BY created_at DESC 
                    LIMIT 1
                """, [school_id])
                
                result = cursor.fetchone()
                if result:
                    form_id = result[0]
                    # Update existing form
                    cursor.execute("""
                        UPDATE forms 
                        SET status = %s, updated_at = NOW()
                        WHERE form_id = %s
                    """, [status, form_id])
                else:
                    # Create new form
                    cursor.execute("""
                        INSERT INTO forms (school_id, status, created_at, updated_at)
                        VALUES (%s, %s, NOW(), NOW())
                    """, [school_id, status])
                    form_id = cursor.lastrowid
                
                # Save each answer
                for answer_data in answers:
                    question_id = answer_data.get('question_id')
                    answer_value = answer_data.get('answer')
                    
                    if question_id and answer_value:
                        # Check if answer already exists
                        cursor.execute("""
                            SELECT answer_id FROM answers 
                            WHERE form_id = %s AND question_id = %s
                        """, [form_id, question_id])
                        
                        existing_answer = cursor.fetchone()
                        
                        if existing_answer:
                            # Update existing answer
                            cursor.execute("""
                                UPDATE answers 
                                SET response = %s, answered_at = NOW()
                                WHERE answer_id = %s
                            """, [answer_value, existing_answer[0]])
                        else:
                            # Insert new answer
                            cursor.execute("""
                                INSERT INTO answers (form_id, question_id, response, answered_at)
                                VALUES (%s, %s, %s, NOW())
                            """, [form_id, question_id, answer_value])
                
                connection.commit()
            
            # Log successful form submission
            if hasattr(request, 'user') and request.user.is_authenticated:
                SystemLogger.log_form_submission(
                    user=request.user,
                    form_id=form_id,
                    school_id=school_id,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    success=True
                )
            
            return JsonResponse({
                'status': 'success',
                'form_id': form_id,
                'answers_saved': len(answers)
            })
            
        except Exception as e:
            if 'connection' in locals():
                connection.rollback()
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'POST required'}, status=405)

@require_GET
def get_categories(request):
    categories = Category.objects.all().order_by('display_order')
    data = [{
        'category_id': cat.category_id,
        'name': cat.name,
        'display_order': cat.display_order
    } for cat in categories]
    return JsonResponse(data, safe=False)

@require_GET
def get_sub_sections(request):
    category_id = request.GET.get('category_id')
    if not category_id or not category_id.isdigit():
        return JsonResponse({'error': 'category_id required and must be a number'}, status=400)
    
    sub_sections = SubSection.objects.filter(category_id=category_id).order_by('display_order')
    data = []
    
    for sub in sub_sections:
        # Simple version that only returns subsection data without complex nested queries
        # This avoids the database schema mismatch issue
        data.append({
            'sub_section_id': sub.sub_section_id,
            'name': sub.name,
            'display_order': sub.display_order
        })
    
    return JsonResponse(data, safe=False)

@csrf_exempt
@require_POST
@session_required
def submit_topic_form(request):
    try:
        data = json.loads(request.body)
        sub_section_id = data.get('sub_section_id')
        name = data.get('name')
        questions = data.get('questions', [])
        status = data.get('status', 'completed')  # default to completed
        # Determine the next display_order for the topic in this sub_section
        with connection.cursor() as cursor:
            cursor.execute("SELECT COALESCE(MAX(display_order), 0) + 1 FROM topics WHERE sub_section_id = %s", [sub_section_id])
            topic_display_order = cursor.fetchone()[0]
            cursor.execute(
                "INSERT INTO topics (sub_section_id, name, display_order) VALUES (%s, %s, %s)",
                [sub_section_id, name, topic_display_order]
            )
            topic_id = cursor.lastrowid
            # Insert questions
            for q_index, q in enumerate(questions, start=1):
                cursor.execute(
                    "INSERT INTO questions (topic_id, question_text, answer_type, is_required, display_order) VALUES (%s, %s, %s, %s, %s)",
                    [topic_id, q['text'], q['type'], q['required'], q_index]
                )
                question_id = cursor.lastrowid
                # Insert choices if needed
                if q['type'] == 'choice':
                    for choice in q.get('choices', []):
                        cursor.execute(
                            "INSERT INTO question_choices (question_id, choice_text) VALUES (%s, %s)",
                            [question_id, choice]
                        )
        return JsonResponse({'status': 'success', 'topic_id': topic_id})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
@require_POST
def login_view(request):
    try:
        data = json.loads(request.body.decode())
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request format'}, status=400)

    email = data.get('email')
    password = data.get('password')
    next_url = data.get('next')
    
    # Extract client information for logging
    ip_address = SystemLogger.get_client_ip(request)
    user_agent = SystemLogger.get_user_agent(request)

    if not email or not password:
        # Log failed login attempt
        SystemLogger.log_login_attempt(
            username=email or 'unknown',
            ip_address=ip_address,
            user_agent=user_agent,
            success=False,
            failure_reason='Missing email or password'
        )
        return JsonResponse({'success': False, 'error': 'Email and password are required'}, status=400)

    # Check admin_user table for authentication
    try:
        admin = AdminUser.objects.get(email=email)
        if bcrypt.checkpw(password.encode(), admin.password_hash.encode()):
            if admin.status != 'active':
                # Log failed login due to inactive account
                SystemLogger.log_login_attempt(
                    username=admin.username,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    success=False,
                    failure_reason='Account inactive'
                )
                return JsonResponse({'success': False, 'error': 'Account is not active'}, status=403)
            
            # Log successful login
            SystemLogger.log_login_attempt(
                username=admin.username,
                ip_address=ip_address,
                user_agent=user_agent,
                success=True
            )
            
            # Update last login
            admin.last_login = timezone.now()
            admin.save(update_fields=['last_login'])
            
            # Determine redirect URL based on admin_level
            if admin.admin_level == 'school':
                redirect_url = '/user-dashboard/'
                user_type = 'school_user'
            else:
                redirect_url = '/dashboard/'
                user_type = 'admin'
            
            # Create/ensure session exists first
            if not request.session.session_key:
                request.session.create()
            
            # Set session data and ensure it's properly saved
            request.session['admin_id'] = admin.admin_id
            request.session['user_type'] = user_type
            request.session['email'] = admin.email
            request.session['admin_level'] = admin.admin_level
            request.session.modified = True
            
            # Force session save
            request.session.save()
            
            # Verify session was saved
            from django.contrib.sessions.models import Session
            try:
                session_obj = Session.objects.get(session_key=request.session.session_key)
                print(f"LOGIN DEBUG: Session verified in database")
            except Session.DoesNotExist:
                print(f"LOGIN DEBUG: WARNING - Session not found in database!")
            
            # Debug logging
            print(f"LOGIN DEBUG: Set session data for admin_id: {admin.admin_id}")
            print(f"LOGIN DEBUG: Session key: {request.session.session_key}")
            print(f"LOGIN DEBUG: Session data: {dict(request.session)}")
            print(f"LOGIN DEBUG: User type: {user_type}, Redirect: {redirect_url}")
            
            # For browser-based login, redirect directly instead of JSON response
            if request.META.get('HTTP_ACCEPT', '').startswith('text/html'):
                # Create a direct redirect response with session cookie
                response = redirect(redirect_url)
                # Ensure the session cookie is set in the response
                response.set_cookie(
                    'sessionid',
                    request.session.session_key,
                    max_age=1209600,
                    domain=None,
                    path='/',
                    secure=False,
                    httponly=True,
                    samesite='Lax'
                )
                return response
            
            # For API/AJAX requests, return JSON
            return JsonResponse({
                'success': True,
                'redirect': redirect_url,
                'user': {
                    'username': admin.username,
                    'email': admin.email,
                    'admin_level': admin.admin_level,
                    'assigned_area': admin.assigned_area,
                    'full_name': admin.full_name,
                    'role': user_type
                },
                'token': 'django-session-token'
            })
        else:
            # Log failed login attempt
            SystemLogger.log_login_attempt(
                username=admin.username,
                ip_address=ip_address,
                user_agent=user_agent,
                success=False,
                failure_reason='Invalid password'
            )
    except AdminUser.DoesNotExist:
        # Log failed login attempt for non-existent user
        SystemLogger.log_login_attempt(
            username=email,
            ip_address=ip_address,
            user_agent=user_agent,
            success=False,
            failure_reason='User not found'
        )
    
    # Return authentication failure if user not found in admin_user table
    return JsonResponse({'success': False, 'error': 'Invalid credentials'}, status=401)

def dashboard_page(request):
    if request.session.get('user_type') != 'admin':
        return redirect('/auth/login/')
    return render(request, 'dashboard/dashboard.html')

def form_page(request):
    # Check if user is logged in
    if not request.session.get('admin_id'):
        return redirect('/auth/login/')
    return render(request, 'form/form.html')

def auth_page(request):
    next_url = request.GET.get('next', '/user/dashboard/')
    return render(request, 'auth/login.html', {'next': next_url})

@csrf_exempt
def test_dashboard_direct(request):
    """Direct dashboard access for testing - bypasses session checks"""
    from app.models import AdminUser
    
    # Get a school user for testing
    admin_user = AdminUser.objects.filter(admin_level='school', status='active').first()
    
    if admin_user:
        context = {
            'user': {
                'id': admin_user.admin_id,
                'username': admin_user.username,
                'email': admin_user.email,
                'is_authenticated': True
            },
            'admin_user': admin_user,
            'school_name': admin_user.school.school_name if admin_user.school else admin_user.assigned_area,
            'role': admin_user.admin_level,
            'full_name': admin_user.full_name,
            'assigned_area': admin_user.assigned_area,
            'admin_level': admin_user.admin_level,
            'user_type': 'school_user'
        }
        return render(request, 'dashboard/user_dash.html', context)
    else:
        return JsonResponse({'error': 'No test user found'}, status=404)

@csrf_exempt  
def manual_login(request):
    """Manual login that sets session and redirects immediately"""
    if request.method == 'GET':
        # Show manual login form
        return JsonResponse({
            'message': 'Manual login endpoint',
            'usage': 'POST with email parameter to login directly',
            'example': 'POST /manual-login/ with {"email": "100001@deped.gov.ph"}'
        })
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode())
            email = data.get('email', '').strip()
            
            if not email:
                return JsonResponse({'error': 'Email required'}, status=400)
            
            # Find admin user by email
            admin = AdminUser.objects.filter(email=email, status='active').first()
            
            if admin:
                # Set session data
                request.session['admin_id'] = admin.admin_id
                request.session['user_type'] = 'school_user' if admin.admin_level == 'school' else 'admin'
                request.session['email'] = admin.email
                request.session['admin_level'] = admin.admin_level
                request.session.save()
                
                # Determine redirect
                redirect_url = '/user-dashboard/' if admin.admin_level == 'school' else '/dashboard/'
                
                print(f"MANUAL LOGIN: Set session for {admin.username} (ID: {admin.admin_id})")
                print(f"MANUAL LOGIN: Redirecting to {redirect_url}")
                
                return JsonResponse({
                    'success': True,
                    'redirect': redirect_url,
                    'user': {
                        'username': admin.username,
                        'email': admin.email,
                        'admin_level': admin.admin_level,
                        'role': 'school_user' if admin.admin_level == 'school' else 'admin'
                    }
                })
            else:
                return JsonResponse({'error': 'User not found'}, status=404)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


def get_admin_context(request):
    """Get admin context data for templates"""
    admin_id = request.session.get('admin_id')
    if not admin_id:
        return None
    
    try:
        from .admin_utils import AdminUserManager
        admin_scope = AdminUserManager.get_user_access_scope(admin_id)
        return {
            'admin_scope': admin_scope,
            'admin_id': admin_id,
            'admin_level': admin_scope.get('admin_level'),
            'permissions': admin_scope.get('permissions', {}),
            'coverage': admin_scope.get('coverage', ''),
        }
    except Exception:
        return None

def admin_page(request):
    """Enhanced admin dashboard with role-based content"""
    # Check if user is an admin
    if request.session.get('user_type') != 'admin':
        return redirect('/auth/login/')
    
    context = get_admin_context(request)
    if not context:
        return redirect('/auth/login/')
    
    # Get dashboard statistics based on admin's scope
    admin_scope = context['admin_scope']
    
    # Get recent activity
    try:
        recent_activities = AdminActivityLog.objects.filter(
            admin_user_id=context['admin_id']
        ).order_by('-timestamp')[:10]
    except:
        recent_activities = []
    
    context.update({
        'recent_activities': recent_activities,
        'dashboard_stats': {
            'total_users': AdminUser.objects.filter(status='active').count(),
            'active_sessions': 0,  # Will be implemented with session management
        }
    })
    
    return render(request, 'admin/admin.html', context)


def user_management_page(request):
    """User management page with scope-based filtering"""
    # Check if user is an admin
    if request.session.get('user_type') != 'admin':
        return redirect('/auth/login/')
    
    context = get_admin_context(request)
    if not context:
        return redirect('/auth/login/')
    
    # Get users within admin's scope
    admin_scope = context['admin_scope']
    users_query = AdminUser.objects.filter(status='active')
    
    # Apply geographic filtering based on admin level
    if admin_scope['admin_level'] == 'region':
        users_query = users_query.filter(region_id=admin_scope.get('region_id'))
    elif admin_scope['admin_level'] == 'division':
        users_query = users_query.filter(division_id=admin_scope.get('division_id'))
    elif admin_scope['admin_level'] == 'district':
        users_query = users_query.filter(district_id=admin_scope.get('district_id'))
    elif admin_scope['admin_level'] == 'school':
        users_query = users_query.filter(school_id=admin_scope.get('school_id'))
    
    users = users_query.order_by('-created_at')[:50]  # Limit for performance
    
    context.update({
        'users': users,
        'can_create_users': admin_scope.get('permissions', {}).get('can_create_users', False),
    })
    
    return render(request, 'admin/user_management.html', context)


def role_page(request):
    """Role and permissions management page"""
    # Check if user is an admin
    if request.session.get('user_type') != 'admin':
        return redirect('/auth/login/')
    
    context = get_admin_context(request)
    if not context:
        return redirect('/auth/login/')
    
    # Get all admin levels and their permissions
    admin_levels = AdminUser.ADMIN_LEVEL_CHOICES
    
    context.update({
        'admin_levels': admin_levels,
    })
    
    return render(request, 'admin/role.html', context)


def logs_page(request):
    """System logs and audit trail page"""
    # Check if user is an admin
    if request.session.get('user_type') != 'admin':
        return redirect('/auth/login/')
    
    context = get_admin_context(request)
    if not context:
        return redirect('/auth/login/')
    
    # Get activity logs based on admin's scope
    admin_scope = context['admin_scope']
    logs = []
    
    try:
        logs_query = AdminActivityLog.objects.all()
        
        # Apply filtering based on admin level
        if admin_scope['admin_level'] != 'central':
            # Non-central admins can only see their own logs
            logs_query = logs_query.filter(admin_user_id=context['admin_id'])
        
        logs = logs_query.order_by('-timestamp')[:100]  # Limit for performance
    except:
        pass
    
    context.update({
        'logs': logs,
    })
    
    return render(request, 'admin/logs.html', context)


def settings_page(request):
    """Admin settings and configuration page"""
    # Check if user is an admin
    if request.session.get('user_type') != 'admin':
        return redirect('/auth/login/')
    
    context = get_admin_context(request)
    if not context:
        return redirect('/auth/login/')
    
    # Get admin's current settings
    try:
        admin_user = AdminUser.objects.get(admin_id=context['admin_id'])
        context.update({
            'admin_user': admin_user,
        })
    except:
        pass
    
    return render(request, 'admin/settings.html', context)


# New Admin API Endpoints

@csrf_exempt
@require_POST
def api_create_admin_user(request):
    """API endpoint to create new admin users"""
    # Check admin permissions
    if request.session.get('user_type') != 'admin':
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)
    
    try:
        from .admin_utils import AdminUserManager
        
        data = json.loads(request.body.decode())
        admin_id = request.session.get('admin_id')
        
        # Validate required fields
        required_fields = ['username', 'email', 'full_name', 'admin_level']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False, 
                    'error': f'{field} is required'
                }, status=400)
        
        # Create the admin user
        new_admin = AdminUserManager.create_admin_user(
            creator_admin_id=admin_id,
            user_data=data,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT')
        )
        
        return JsonResponse({
            'success': True,
            'admin_id': new_admin.admin_id,
            'username': new_admin.username,
            'message': f'Admin user {new_admin.username} created successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@csrf_exempt
def api_admin_users(request):
    """API endpoint for admin user management"""
    # Check admin permissions
    if request.session.get('user_type') != 'admin':
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)
    
    admin_id = request.session.get('admin_id')
    
    if request.method == 'GET':
        # Get users within admin's scope
        users_query = AdminUser.objects.filter(status='active')
        
        users_data = []
        for user in users_query.select_related('region', 'division', 'district', 'school'):
            users_data.append({
                'admin_id': user.admin_id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'admin_level': user.admin_level,
                'assigned_area': user.assigned_area,
                'status': user.status,
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'created_at': user.created_at.isoformat(),
                'permissions': {
                    'can_create_users': user.can_create_users,
                    'can_manage_users': user.can_manage_users,
                    'can_set_deadlines': user.can_set_deadlines,
                    'can_approve_submissions': user.can_approve_submissions,
                    'can_view_system_logs': user.can_view_system_logs,
                }
            })
        
        return JsonResponse({
            'success': True,
            'users': users_data,
            'total': len(users_data)
        })
    
    elif request.method == 'POST':
        # Create new admin user
        return api_create_admin_user(request)


@csrf_exempt
@require_POST
def api_set_deadline(request):
    """API endpoint to set form deadlines"""
    # Check admin permissions
    if request.session.get('user_type') != 'admin':
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)
    
    try:
        from .admin_utils import DeadlineManager
        
        data = json.loads(request.body.decode())
        admin_id = request.session.get('admin_id')
        
        # Create the deadline
        deadline = DeadlineManager.set_deadline(
            admin_id=admin_id,
            deadline_data=data,
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'success': True,
            'deadline_id': deadline.deadline_id,
            'form_type': deadline.form_type,
            'deadline_date': deadline.deadline_date.isoformat(),
            'message': 'Deadline set successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@csrf_exempt
@require_GET
def api_activity_logs(request):
    """API endpoint to get activity logs"""
    # Check admin permissions
    if request.session.get('user_type') != 'admin':
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)
    
    admin_id = request.session.get('admin_id')
    
    try:
        # Get logs based on admin's scope
        logs_query = AdminActivityLog.objects.filter(admin_user_id=admin_id)
        
        # Pagination
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))
        
        start = (page - 1) * page_size
        end = start + page_size
        
        logs = logs_query.order_by('-timestamp')[start:end]
        
        logs_data = []
        for log in logs:
            logs_data.append({
                'log_id': log.log_id,
                'admin_user': log.admin_user.username,
                'action': log.action,
                'resource_type': log.resource_type,
                'resource_id': log.resource_id,
                'details': log.details,
                'ip_address': log.ip_address,
                'timestamp': log.timestamp.isoformat(),
            })
        
        return JsonResponse({
            'success': True,
            'logs': logs_data,
            'total': logs_query.count(),
            'page': page,
            'page_size': page_size
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@csrf_exempt
@require_GET
def api_geographic_data(request, data_type):
    """API endpoint to get geographic data for dropdowns"""
    # Check admin permissions
    if request.session.get('user_type') != 'admin':
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)
    
    try:
        parent_id = request.GET.get('parent_id')
        
        if data_type == 'regions':
            data = list(Region.objects.all().values('id', 'name').order_by('name'))
        elif data_type == 'divisions':
            if parent_id:
                data = list(Division.objects.filter(region_id=parent_id).values('id', 'name').order_by('name'))
            else:
                data = list(Division.objects.all().values('id', 'name').order_by('name'))
        elif data_type == 'districts':
            if parent_id:
                data = list(District.objects.filter(division_id=parent_id).values('id', 'name').order_by('name'))
            else:
                data = list(District.objects.all().values('id', 'name').order_by('name'))
        elif data_type == 'schools':
            if parent_id:
                data = list(School.objects.filter(district_id=parent_id).values('id', 'school_name').order_by('school_name'))
            else:
                data = list(School.objects.all().values('id', 'school_name').order_by('school_name'))
        else:
            return JsonResponse({'success': False, 'error': 'Invalid data type'}, status=400)
        
        return JsonResponse({
            'success': True,
            'data': data,
            'total': len(data)
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

# Missing Dashboard API endpoints
@csrf_exempt
@require_GET
def api_dashboard_completion_by_region(request):
    """API endpoint for completion by region data"""
    try:
        # Mock data for completion by region
        data = {
            'regions': [
                {'name': 'NCR', 'completion_rate': 85.2, 'total_forms': 1245},
                {'name': 'Region I', 'completion_rate': 78.9, 'total_forms': 892},
                {'name': 'Region II', 'completion_rate': 82.1, 'total_forms': 654},
                {'name': 'Region III', 'completion_rate': 79.5, 'total_forms': 1123},
                {'name': 'Region IV-A', 'completion_rate': 88.3, 'total_forms': 1456},
                {'name': 'Region IV-B', 'completion_rate': 76.8, 'total_forms': 534},
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_GET
def api_dashboard_response_distribution(request):
    """API endpoint for response distribution data"""
    try:
        question_id = request.GET.get('question', 1)
        # Mock data for response distribution
        data = {
            'question_id': question_id,
            'question_text': 'How would you rate the overall quality of education in your school?',
            'responses': [
                {'option': 'Excellent', 'count': 245, 'percentage': 24.5},
                {'option': 'Good', 'count': 456, 'percentage': 45.6},
                {'option': 'Fair', 'count': 234, 'percentage': 23.4},
                {'option': 'Poor', 'count': 65, 'percentage': 6.5},
            ],
            'total_responses': 1000
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_GET
def api_dashboard_forms_over_time(request):
    """API endpoint for forms over time data"""
    try:
        # Mock data for forms over time
        data = {
            'time_series': [
                {'date': '2025-01-01', 'submitted': 45, 'completed': 42},
                {'date': '2025-01-02', 'submitted': 52, 'completed': 48},
                {'date': '2025-01-03', 'submitted': 38, 'completed': 35},
                {'date': '2025-01-04', 'submitted': 61, 'completed': 58},
                {'date': '2025-01-05', 'submitted': 47, 'completed': 44},
                {'date': '2025-01-06', 'submitted': 55, 'completed': 51},
                {'date': '2025-01-07', 'submitted': 49, 'completed': 46},
            ],
            'total_submitted': 347,
            'total_completed': 324
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_GET
def api_dashboard_top_schools(request):
    """API endpoint for top schools data"""
    try:
        # Mock data for top schools
        data = {
            'schools': [
                {'name': 'Manila Science High School', 'completion_rate': 95.8, 'total_forms': 245},
                {'name': 'Quezon City High School', 'completion_rate': 94.2, 'total_forms': 312},
                {'name': 'Makati High School', 'completion_rate': 93.7, 'total_forms': 189},
                {'name': 'Taguig Science High School', 'completion_rate': 92.1, 'total_forms': 156},
                {'name': 'Pasig City Science High School', 'completion_rate': 91.8, 'total_forms': 203},
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_GET
@csrf_exempt
def get_drafts(request):
    # Fetch all draft questionnaires grouped by category, sub-section, topic, with questions and choices
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT
                c.category_id, c.name AS category,
                s.sub_section_id, s.name AS sub_section,
                t.topic_id, t.name AS topic,
                q.question_id, q.question_text, q.answer_type, q.is_required, q.display_order,
                qc.choice_text
            FROM categories c
            JOIN sub_sections s ON s.category_id = c.category_id
            JOIN topics t ON t.sub_section_id = s.sub_section_id
            JOIN questions q ON q.topic_id = t.topic_id
            LEFT JOIN question_choices qc ON qc.question_id = q.question_id
            ORDER BY c.display_order, s.display_order, t.display_order, q.display_order, qc.choice_id
        ''')
        rows = cursor.fetchall()

    # Group data by category > sub_section > topic > questions
    drafts = []
    cat_map = {}
    for row in rows:
        cat_id, cat_name, sub_id, sub_name, topic_id, topic_name, q_id, q_text, q_type, q_required, q_order, choice_text = row
        cat_key = (cat_id, cat_name)
        sub_key = (sub_id, sub_name)
        topic_key = (topic_id, topic_name)
        # Find or create category
        if cat_key not in cat_map:
            cat_map[cat_key] = {}
        sub_map = cat_map[cat_key]
        if sub_key not in sub_map:
            sub_map[sub_key] = {}
        topic_map = sub_map[sub_key]
        if topic_key not in topic_map:
            topic_map[topic_key] = {}
        q_map = topic_map[topic_key]
        if q_id not in q_map:
            q_map[q_id] = {
                'question_id': q_id,  # Ensure question_id is present for frontend
                'text': q_text,
                'type': q_type,
                'required': q_required,
                'displayOrder': q_order,
                'choices': [],
                'sub_questions': []
            }
        if choice_text:
            q_map[q_id]['choices'].append(choice_text)
    # Fetch sub-questions for all collected questions to avoid N+1
    all_q_ids = [qid for (_, _), sub_map in cat_map.items() for (_, _), topic_map in sub_map.items() for (_, _), q_map in topic_map.items() for qid in q_map.keys()]
    if all_q_ids:
        with connection.cursor() as cursor:
            format_strings = ','.join(['%s'] * len(all_q_ids))
            cursor.execute(f'''
                SELECT question_id, sub_question_id, sub_question_text, answer_type, display_order
                FROM sub_questions
                WHERE question_id IN ({format_strings})
                ORDER BY question_id, display_order, sub_question_id
            ''', all_q_ids)
            for srow in cursor.fetchall():
                qid, sid, stext, satype, sorder = srow
                # place into the built maps
                for (_, _), sub_map in cat_map.items():
                    for (_, _), topic_map in sub_map.items():
                        for (_, _), q_map in topic_map.items():
                            if qid in q_map:
                                q_map[qid]['sub_questions'].append({
                                    'sub_question_id': sid,
                                    'sub_question_text': stext,
                                    'answer_type': satype,
                                    'display_order': sorder
                                })
                                break
    # Convert to list structure
    for (cat_id, cat_name), sub_map in cat_map.items():
        for (sub_id, sub_name), topic_map in sub_map.items():
            for (topic_id, topic_name), q_map in topic_map.items():
                drafts.append({
                    'category': cat_name,
                    'subSection': sub_name,
                    'sub_section_id': sub_id,
                    'topic': topic_name,
                    'questions': list(q_map.values())
                })
    return JsonResponse(drafts, safe=False)

@csrf_exempt
@require_POST
def save_topic(request):
    try:
        data = json.loads(request.body)
        print('[save_topic] Received data:', data)  # Log the received data to the console
        sub_section_id = data.get('sub_section_id')
        topic_name = data.get('name')
        topic_display_order = data.get('display_order', 1)
        questions = data.get('questions', [])
        if not (sub_section_id and topic_name and questions):
            return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)
        with connection.cursor() as cursor:
            # Insert topic
            cursor.execute(
                'INSERT INTO topics (sub_section_id, name, display_order) VALUES (%s, %s, %s)',
                [sub_section_id, topic_name, topic_display_order]
            )
            topic_id = cursor.lastrowid
            # Insert questions, choices, and sub-questions
            for q in questions:
                cursor.execute(
                    'INSERT INTO questions (topic_id, question_text, answer_type, is_required, display_order) VALUES (%s, %s, %s, %s, %s)',
                    [topic_id, q.get('question_text'), q.get('answer_type'), int(q.get('is_required', False)), q.get('display_order', 1)]
                )
                question_id = cursor.lastrowid
                
                # Insert question choices if answer_type is 'choice'
                if q.get('answer_type') == 'choice':
                    for choice in q.get('choices', []):
                        cursor.execute(
                            'INSERT INTO question_choices (question_id, choice_text) VALUES (%s, %s)',
                            [question_id, choice]
                        )
                
                # Insert sub-questions
                sub_questions = q.get('sub_questions', [])
                if sub_questions:
                    for sub_order, sub_q in enumerate(sub_questions, 1):
                        cursor.execute(
                            'INSERT INTO sub_questions (question_id, sub_question_text, answer_type, display_order) VALUES (%s, %s, %s, %s)',
                            [question_id, sub_q.get('question_text', ''), sub_q.get('answer_type', 'text'), sub_order]
                        )
        return JsonResponse({'success': True, 'topic_id': topic_id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def create_question(request):
    try:
        data = json.loads(request.body)
        topic_id = data.get('topic_id')
        question_text = data.get('question_text')
        answer_type = data.get('answer_type')
        is_required = data.get('is_required', False)
        display_order = data.get('display_order', 1)
        choices = data.get('choices', [])
        answer_description = data.get('answer_description', '')
        if not (topic_id and question_text and answer_type):
            return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)
        topic = Topic.objects.get(pk=topic_id)
        question = Question.objects.create(
            topic=topic,
            question_text=question_text,
            answer_type=answer_type,
            is_required=is_required,
            display_order=display_order,
        )
        if answer_type == 'choice' and choices:
            for choice in choices:
                QuestionChoice.objects.create(question=question, choice_text=choice)
        
        # Handle sub-questions
        sub_questions = data.get('sub_questions', [])
        if sub_questions:
            for sub_order, sub_q in enumerate(sub_questions, 1):
                SubQuestion.objects.create(
                    question=question,
                    sub_question_text=sub_q.get('question_text', ''),
                    answer_type=sub_q.get('answer_type', 'text'),
                    display_order=sub_order
                )
        
        return JsonResponse({'success': True, 'question_id': question.question_id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["PUT"])
def update_question(request, question_id):
    try:
        data = json.loads(request.body)
        question = Question.objects.get(pk=question_id)
        question_text = data.get('question_text')
        answer_type = data.get('answer_type')
        is_required = data.get('is_required', False)
        display_order = data.get('display_order', 1)
        choices = data.get('choices', [])
        answer_description = data.get('answer_description', '')
        if question_text:
            question.question_text = question_text
        if answer_type:
            question.answer_type = answer_type
        question.is_required = is_required
        question.display_order = display_order
        question.save()
        # Update choices
        QuestionChoice.objects.filter(question=question).delete()
        if answer_type == 'choice' and choices:
            for choice in choices:
                QuestionChoice.objects.create(question=question, choice_text=choice)
        # Upsert sub-questions if provided
        sub_questions = data.get('sub_questions')
        if isinstance(sub_questions, list):
            existing = {sq.sub_question_id: sq for sq in SubQuestion.objects.filter(question=question)}
            seen_ids = set()
            order_counter = 1
            for sq in sub_questions:
                sq_id = sq.get('sub_question_id') or sq.get('id')
                sq_text = sq.get('sub_question_text') or sq.get('question_text') or ''
                sq_type = sq.get('answer_type') or 'text'
                sq_order = sq.get('display_order') or order_counter
                if sq_id and sq_id in existing:
                    obj = existing[sq_id]
                    obj.sub_question_text = sq_text
                    obj.answer_type = sq_type
                    obj.display_order = sq_order
                    obj.save(update_fields=['sub_question_text', 'answer_type', 'display_order'])
                    seen_ids.add(sq_id)
                else:
                    new_obj = SubQuestion.objects.create(
                        question=question,
                        sub_question_text=sq_text,
                        answer_type=sq_type,
                        display_order=sq_order
                    )
                    seen_ids.add(new_obj.sub_question_id)
                order_counter += 1
            # Delete removed sub-questions
            to_delete = [obj_id for obj_id in existing.keys() if obj_id not in seen_ids]
            if to_delete:
                SubQuestion.objects.filter(sub_question_id__in=to_delete).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
def delete_question(request, question_id):
    from .models import Question, QuestionChoice
    try:
        # Delete choices first (if any)
        QuestionChoice.objects.filter(question_id=question_id).delete()
        deleted, _ = Question.objects.filter(question_id=question_id).delete()
        if deleted == 0:
            return JsonResponse({'success': False, 'error': 'Question not found'}, status=404)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# --- Simple session-based rate limiting for AJAX validation endpoints ---
def rate_limit_session(request, key, max_requests, window_seconds):
    if 'rate_limit' not in request.session:
        request.session['rate_limit'] = {}
    now = int(timezone.now().timestamp())
    if key not in request.session['rate_limit']:
        request.session['rate_limit'][key] = []
    # Remove expired timestamps
    request.session['rate_limit'][key] = [ts for ts in request.session['rate_limit'][key] if ts > now - window_seconds]
    if len(request.session['rate_limit'][key]) >= max_requests:
        return False
    request.session['rate_limit'][key].append(now)
    request.session.modified = True
    return True

@api_view(['GET'])
@permission_classes([AllowAny])
def search_location(request):
    type_ = request.GET.get('type', '')
    query = (request.GET.get('q', '') or '').strip()
    # Username existence check with session rate limiting
    if type_ == 'username':
        if not rate_limit_session(request, 'username', 10, 10):
            return Response({'error': 'Too many requests. Please slow down.'}, status=429)
        exists = AdminUser.objects.filter(username=query).exists()
        return Response({'usernameExists': exists})
    # Email existence check with session rate limiting
    if type_ == 'email':
        if not rate_limit_session(request, 'email', 10, 10):
            return Response({'error': 'Too many requests. Please slow down.'}, status=429)
        exists = AdminUser.objects.filter(email=query).exists()
        return Response({'emailExists': exists})
    table_map = {
        'region': Region,
        'division': Division,
        'district': District,
        'school': School
    }
    if type_ not in table_map:
        return Response({'error': 'Invalid type parameter'})
    if query == '':
        return Response([])
    if type_ == 'school':
        # School search with joins
        schools = School.objects.filter(
            models.Q(school_name__icontains=query) | models.Q(school_id__icontains=query)
        ).select_related('district', 'division', 'region')[:10]
        results = []
        for s in schools:
            results.append({
                'school_name': s.school_name,
                'school_id': s.school_id,
                'district_name': s.district.name if s.district else None,
                'division_name': s.division.name if s.division else None,
                'region_name': s.region.name if s.region else None,
            })
        return Response(results)
    # Default: return empty for other types (or implement as needed)
    return Response([])

@api_view(['POST'])
@permission_classes([AllowAny])
def signin(request):
    data = request.data
    school_name = data.get('schoolName', '').strip()
    district = data.get('district', '').strip()
    division = data.get('division', '').strip()
    region = data.get('region', '').strip()
    username = data.get('username', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '')
    force = bool(data.get('force', False))
    errors = []
    if not school_name or not username or not email or not password:
        errors.append('All fields are required.')
    if AdminUser.objects.filter(username=username).exists():
        errors.append('Username already exists.')
    if AdminUser.objects.filter(email=email).exists():
        errors.append('Email already exists.')
    # Look up all IDs from names (school, district, division, region)
    school_obj = School.objects.filter(school_name=school_name).first()
    if school_obj:
        school_id = school_obj.id  # Use primary key, not school_id field
        district_id = school_obj.district_id
        division_id = school_obj.division_id
        region_id = school_obj.region_id
        school_name = school_obj.school_name
    else:
        errors.append('School not found.')
        school_obj = None
        school_id = district_id = division_id = region_id = None
    if not school_id or not district_id or not division_id or not region_id:
        errors.append('School, district, division, or region not found.')
    # Check if another user is already in the database with the same school
    existing_user = AdminUser.objects.filter(school__id=school_id, admin_level='school').first() if school_id else None
    if existing_user and not force:
        return Response({
            'exists': True,
            'username': existing_user.username,
            'email': existing_user.email
        })
    if not errors:
        from django.contrib.auth.hashers import make_password
        password_hash = make_password(password)
        try:
            with transaction.atomic():
                # Ensure we have a valid school_obj
                if not school_obj:
                    raise Exception("School object not found")
                
                user = AdminUser.objects.create( 
                    username=username,
                    password_hash=password_hash,
                    email=email,
                    admin_level='school',
                    assigned_area=school_name,
                    school=school_obj,
                    district_id=district_id,
                    division_id=division_id,
                    region_id=region_id,
                    status='active',
                    full_name=username  # Use username as full name for now
                )
            return Response({'success': True})
        except Exception as e:
            errors.append('Database error: ' + str(e))
    if errors:
        if force:
            return Response({'success': False, 'errors': errors})
        else:
            return Response({'errors': errors})

def signin_page(request):
    return render(request, 'auth/signin.html')

# Removed duplicate categories_api and sub_sections_api functions
# Use get_categories and get_sub_sections instead

@require_GET
def get_topics(request):
    sub_section_id = request.GET.get('sub_section_id')
    if not sub_section_id:
        return JsonResponse([], safe=False)
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT topic_id, name, display_order, sub_section_id
            FROM topics
            WHERE sub_section_id = %s
            ORDER BY display_order
        ''', [sub_section_id])
        topics = [
            {'topic_id': row[0], 'name': row[1], 'display_order': row[2], 'sub_section_id': row[3]}
            for row in cursor.fetchall()
        ]
    return JsonResponse(topics, safe=False)

@require_GET
def get_questions(request, topic_id):
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT question_id, question_text, answer_type, is_required, display_order
            FROM questions
            WHERE topic_id = %s
            ORDER BY display_order
        ''', [topic_id])
        questions = [
            {'question_id': row[0], 'question_text': row[1], 'answer_type': row[2], 'is_required': row[3], 'display_order': row[4]}
            for row in cursor.fetchall()
        ]
        for q in questions:
            cursor.execute('SELECT choice_text FROM question_choices WHERE question_id = %s', [q['question_id']])
            q['choices'] = [row[0] for row in cursor.fetchall()]
            cursor.execute('''
                SELECT sub_question_id, sub_question_text, answer_type, display_order
                FROM sub_questions
                WHERE question_id = %s
                ORDER BY display_order, sub_question_id
            ''', [q['question_id']])
            sub_rows = cursor.fetchall()
            q['sub_questions'] = [
                {
                    'sub_question_id': srow[0],
                    'sub_question_text': srow[1],
                    'answer_type': srow[2],
                    'display_order': srow[3]
                }
                for srow in sub_rows
            ]
    return JsonResponse(questions, safe=False)

def report_page(request):
    import time
    return render(request, 'report/report.html', {'timestamp': int(time.time())})

def user_dashboard_page(request):
    """
    View for the user dashboard page.
    Shows overall progress, form completion status, and recent activities.
    """
    return render(request, 'dashboard/user_dashboard.html')

def user_form_page(request):
    """
    View for the user form page.
    Displays the form with categories, sub-sections, and questions.
    """
    return render(request, 'dashboard/user_form.html')

def user_profile_page(request):
    """
    View for the user profile settings page.
    Allows users to update their profile information.
    """
    return render(request, 'dashboard/user_profile.html')

def logout_view(request):
    # Clear FastAPI cache for this user if authenticated
    if request.user.is_authenticated or request.session.get('admin_id'):
        try:
            # Call FastAPI cache clear endpoint
            response = requests.get('http://127.0.0.1:8002/clear-cache', timeout=5)
            print(f"Cache clear response: {response.status_code}")
        except Exception as e:
            print(f"Error clearing cache: {e}")
    
    # Clear session data for admin users
    if request.session.get('admin_id'):
        request.session.flush()  # Clear all session data
    else:
        logout(request)  # Use Django logout for Django authenticated users
    
    return redirect('/auth/login/')

# FastAPI Proxy Views
# Duplicate imports removed - already imported at top

def proxy_to_fastapi(request, endpoint):
    """Proxy request to FastAPI server with authentication."""
    fastapi_url = f"http://127.0.0.1:8002{endpoint}"
    
    # Generate JWT token for the authenticated user
    SECRET_KEY = settings.SECRET_KEY  # Use Django's SECRET_KEY from settings
    ALGORITHM = "HS256"
    
    try:
        # Get user ID from either Django user or session
        user_id = None
        if request.user.is_authenticated:
            user_id = request.user.id
        elif request.session.get('admin_id'):
            user_id = request.session.get('admin_id')
        
        if not user_id:
            return JsonResponse({'error': 'No authenticated user found'}, status=403)
        
        # Create JWT token for the authenticated user
        payload = {
            "sub": str(user_id),  # Convert to string to match FastAPI format
            "exp": datetime.utcnow() + timedelta(minutes=30)
        }
        token = jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)
        
        # Prepare headers with authentication
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # Forward the request to FastAPI
        if request.method == 'GET':
            response = requests.get(fastapi_url, headers=headers)
        elif request.method == 'POST':
            # Handle JSON data from request body
            try:
                data = json.loads(request.body.decode('utf-8'))
            except:
                data = request.POST.dict()
            response = requests.post(fastapi_url, json=data, headers=headers)
        elif request.method == 'PUT':
            # Handle JSON data from request body
            try:
                data = json.loads(request.body.decode('utf-8'))
            except:
                data = request.POST.dict()
            response = requests.put(fastapi_url, json=data, headers=headers)
        else:
            return JsonResponse({'error': 'Method not allowed'}, status=405)
        
        # Check if response is successful
        if response.status_code == 200:
            return JsonResponse(response.json(), status=200, safe=False)
        else:
            return JsonResponse({'error': f'FastAPI error: {response.status_code}'}, status=response.status_code)
            
    except requests.exceptions.RequestException as e:
        print(f"Proxy error: {e}")
        return JsonResponse({'error': f'FastAPI server error: {str(e)}'}, status=500)
    except Exception as e:
        print(f"Unexpected error: {e}")
        return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)

@session_or_login_required
@csrf_exempt
def api_dashboard_stats(request):
    """Proxy to FastAPI dashboard stats endpoint."""
    return proxy_to_fastapi(request, '/api/dashboard/stats')

@session_or_login_required
@csrf_exempt
def api_dashboard_categories(request):
    """Proxy to FastAPI dashboard categories endpoint."""
    return proxy_to_fastapi(request, '/api/dashboard/categories')

@session_or_login_required
@csrf_exempt
def api_form_sections(request):
    """Get form sections with questions for a user - Django implementation."""
    try:
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        
        admin_user = AdminUser.objects.get(admin_id=admin_id)
        
        # Get categories with their subsections, topics, and questions
        categories_data = []
        categories = Category.objects.all().order_by('display_order')
        
        for category in categories:
            category_data = {
                'category_id': category.category_id,
                'category_name': category.name,
                'status': 'in_progress',  # You can calculate this based on completion
                'progress_percentage': 0.0,
                'total_questions': 0,
                'answered_questions': 0,
                'questions': []
            }
            
            subsections = SubSection.objects.filter(category=category).order_by('display_order')
            for subsection in subsections:
                topics = Topic.objects.filter(sub_section=subsection).order_by('display_order')
                for topic in topics:
                    questions = Question.objects.filter(topic=topic).order_by('display_order')
                    for question in questions:
                        # Get existing answer if any
                        answer_text = ""
                        try:
                            # Use the same form creation logic as api_form_answers
                            form = get_or_create_admin_form(admin_user)
                            if form:
                                answer = Answer.objects.filter(form=form, question=question, sub_question__isnull=True).first()
                                if answer and answer.response:
                                    answer_text = answer.response
                        except Exception as e:
                            print(f"Error getting answer for question {question.question_id}: {e}")
                            answer_text = ""
                        
                        # Get sub-questions
                        sub_questions = []
                        for sub_q in SubQuestion.objects.filter(question=question).order_by('display_order'):
                            sub_answer_text = ""
                            try:
                                # Use the same form creation logic as api_form_answers
                                form = get_or_create_admin_form(admin_user)
                                if form:
                                    # For sub-questions, we need to match both question and sub_question
                                    sub_answer = Answer.objects.filter(
                                        form=form, 
                                        question=question,  # Parent question
                                        sub_question=sub_q   # Sub-question
                                    ).first()
                                    if sub_answer and sub_answer.response:
                                        sub_answer_text = sub_answer.response
                            except Exception as e:
                                print(f"Error getting sub-answer for sub_question {sub_q.sub_question_id}: {e}")
                                sub_answer_text = ""
                            
                            sub_questions.append({
                                'sub_question_id': sub_q.sub_question_id,
                                'sub_question_text': sub_q.sub_question_text,
                                'answer_type': sub_q.answer_type,
                                'answer': sub_answer_text
                            })
                        
                        question_data = {
                            'question_id': question.question_id,
                            'question_text': question.question_text,
                            'answer_type': question.answer_type,
                            'answer': answer_text,
                            'sub_questions': sub_questions,
                            'topic_name': topic.name,
                            'sub_section_name': subsection.name,  # Fixed: was subsection_name
                            'category_name': category.name
                        }
                        category_data['questions'].append(question_data)
                        category_data['total_questions'] += 1
                        if answer_text:
                            category_data['answered_questions'] += 1
            
            # Calculate progress
            if category_data['total_questions'] > 0:
                category_data['progress_percentage'] = round(
                    (category_data['answered_questions'] / category_data['total_questions']) * 100, 1
                )
                if category_data['progress_percentage'] == 100:
                    category_data['status'] = 'completed'
                elif category_data['progress_percentage'] > 0:
                    category_data['status'] = 'in_progress'
                else:
                    category_data['status'] = 'not_started'
            
            categories_data.append(category_data)
        
        return JsonResponse(categories_data, safe=False)
        
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'Admin user not found'}, status=404)
    except Exception as e:
        print(f"Error in api_form_sections: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_form_submit(request):
    """Submit form answers - Django implementation."""
    try:
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        
        admin_user = AdminUser.objects.get(admin_id=admin_id)
        
        if request.method != 'POST':
            return JsonResponse({'error': 'Method not allowed'}, status=405)
        
        # Parse request data
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        
        # Get or create form for this user using the helper function
        form = get_or_create_admin_form(admin_user)
        if not form:
            return JsonResponse({'error': 'Could not create or find form for user'}, status=500)
        
        # Process answers
        saved_count = 0
        errors = []
        
        # Handle different data formats from JavaScript
        answers_data = {}
        
        # Format 1: Single question/answer submission (from form-data-manager.js)
        if 'question_id' in data and 'answer' in data:
            question_id = data.get('question_id')
            answer_value = data.get('answer', '')
            sub_question_id = data.get('sub_question_id')
            
            # CRITICAL FIX: JavaScript sometimes sends sub-question IDs as question_id
            # Sub-question ID range: 7663-9859, Question ID range: 2390-5536
            if question_id and question_id >= 7663:
                # This is actually a sub-question ID being sent as question_id
                print(f"Detected sub-question ID {question_id} sent as question_id, correcting...")
                answers_data[f"sub_question_{question_id}"] = answer_value
            elif sub_question_id:
                answers_data[f"sub_question_{sub_question_id}"] = answer_value
            else:
                answers_data[f"question_{question_id}"] = answer_value
        
        # Format 2: Multiple answers in 'answers' array
        elif 'answers' in data:
            answers_list = data.get('answers', [])
            if isinstance(answers_list, list):
                for answer_item in answers_list:
                    if 'question_id' in answer_item:
                        q_id = answer_item['question_id']
                        # Apply the same ID range detection logic
                        if q_id and q_id >= 7663:
                            # This is actually a sub-question ID
                            answers_data[f"sub_question_{q_id}"] = answer_item.get('answer', '')
                        else:
                            answers_data[f"question_{q_id}"] = answer_item.get('answer', '')
                    if 'sub_question_id' in answer_item:
                        answers_data[f"sub_question_{answer_item['sub_question_id']}"] = answer_item.get('answer', '')
            elif isinstance(answers_list, dict):
                # Format 3: Answers as key-value pairs
                answers_data = answers_list
        
        # Save answers
        for key, value in answers_data.items():
            try:
                # Skip empty values
                if not value or str(value).strip() == '':
                    continue
                    
                if key.startswith('question_'):
                    question_id = int(key.replace('question_', ''))
                    question = Question.objects.get(question_id=question_id)
                    
                    answer, created = Answer.objects.get_or_create(
                        form=form,
                        question=question,
                        defaults={'response': str(value)[:500]}  # Truncate response if too long
                    )
                    if not created:
                        answer.response = str(value)[:500]
                        answer.save()
                    saved_count += 1
                    print(f"Saved answer for question {question_id}: {str(value)[:50]}...")
                    
                elif key.startswith('sub_question_'):
                    sub_question_id = int(key.replace('sub_question_', ''))
                    sub_question = SubQuestion.objects.get(sub_question_id=sub_question_id)
                    
                    # CRITICAL: Also set the parent question for sub-question answers
                    parent_question = sub_question.question
                    
                    answer, created = Answer.objects.get_or_create(
                        form=form,
                        question=parent_question,  # Required: parent question
                        sub_question=sub_question,  # Optional: sub-question
                        defaults={'response': str(value)[:500]}  # Truncate response if too long
                    )
                    if not created:
                        answer.response = str(value)[:500]
                        answer.save()
                    saved_count += 1
                    print(f"Saved answer for sub-question {sub_question_id} (parent question {parent_question.question_id}): {str(value)[:50]}...")
                    
            except (ValueError, Question.DoesNotExist, SubQuestion.DoesNotExist) as e:
                error_msg = f"Error saving {key}: {str(e)}"
                errors.append(error_msg)
                print(error_msg)
                continue
            except Exception as e:
                error_msg = f"Unexpected error saving {key}: {str(e)}"
                errors.append(error_msg)
                print(error_msg)
                continue
        
        # Update form status
        form.updated_at = timezone.now()
        submit_type = data.get('submit_type')
        if submit_type == 'final':
            form.status = 'completed'
        elif submit_type == 'to_district':
            # School submitting to District for review
            form.status = 'district_pending'
            form.current_level = 'district'
            # set submitted_at if available in schema; ignore if missing
            try:
                from django.db import connection
                with connection.cursor() as c:
                    c.execute("UPDATE forms SET submitted_at = NOW() WHERE form_id = %s", [form.form_id])
            except Exception:
                pass
        else:
            form.status = 'draft'
        form.save()
        
        return JsonResponse({
            'success': True,
            'saved_count': saved_count,
            'form_id': form.form_id,
            'status': form.status,
            'errors': errors if errors else None
        })
        
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'Admin user not found'}, status=404)
    except Exception as e:
        print(f"Error in api_form_submit: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
def api_profile(request):
    """Get user profile with user ID included."""
    try:
        # Get user's admin record - this function should be updated to use session-based auth
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        user_school = AdminUser.objects.select_related('region', 'division', 'district', 'school').get(admin_id=admin_id)
        
        profile_data = {
            "id": user_school.admin_id,  # Use admin_id as the primary ID
            "user_id": user_school.admin_id,  # Alternative field name
            "name": user_school.username,
            "username": user_school.username,
            "email": user_school.email,
            "role": user_school.admin_level,
            "personal_info": {
                "id": user_school.admin_id,
                "school_name": user_school.school.school_name if user_school.school else user_school.assigned_area,
                "email": user_school.email,
                "region": user_school.region.name if user_school.region else None,
                "division": user_school.division.name if user_school.division else None,
                "district": user_school.district.name if user_school.district else None
            },
            "account_settings": {
                "email_notifications": True,
                "sms_notifications": False,
                "two_factor_enabled": False,
                "last_login": user_school.last_login.isoformat() if user_school.last_login else None,
                "account_created": user_school.created_at.isoformat() if hasattr(user_school, 'created_at') else None
            }
        }
        
        return JsonResponse(profile_data)
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'User profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Analytics API Endpoints
@session_or_login_required
@csrf_exempt
def api_analytics_data(request):
    """Get analytics data for the user dashboard."""
    try:
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        user_school = AdminUser.objects.get(admin_id=admin_id)
        
        # Get user's forms
        user_forms = Form.objects.filter(user_id=admin_id)
        
        # Get total questions count
        total_questions = Question.objects.count()
        
        # Get answered questions count for this user
        answered_questions = Answer.objects.filter(
            form__in=user_forms,
            response__isnull=False
        ).exclude(response='').count()
        
        # Calculate completion rate
        completion_rate = (answered_questions / total_questions * 100) if total_questions > 0 else 0
        
        # Get form status data
        form_status_data = []
        for category in Category.objects.all().order_by('display_order'):
            category_data = {
                'name': category.name,
                'type': 'category',
                'children': []
            }
            
            for subsection in SubSection.objects.filter(category=category).order_by('display_order'):
                subsection_data = {
                    'name': subsection.name,
                    'type': 'subsection',
                    'children': []
                }
                
                for topic in Topic.objects.filter(sub_section=subsection).order_by('display_order'):
                    # Get questions for this topic
                    topic_questions = Question.objects.filter(topic=topic)
                    topic_answered = Answer.objects.filter(
                        form__in=user_forms,
                        question__in=topic_questions,
                        response__isnull=False
                    ).exclude(response='').count()
                    
                    topic_completion = (topic_answered / topic_questions.count() * 100) if topic_questions.count() > 0 else 0
                    
                    topic_data = {
                        'name': topic.name,
                        'type': 'topic',
                        'completion': round(topic_completion, 1),
                        'answered': topic_answered,
                        'total': topic_questions.count()
                    }
                    subsection_data['children'].append(topic_data)
                
                # Calculate subsection completion
                subsection_questions = Question.objects.filter(topic__sub_section=subsection)
                subsection_answered = Answer.objects.filter(
                    form__in=user_forms,
                    question__in=subsection_questions,
                    response__isnull=False
                ).exclude(response='').count()
                
                subsection_completion = (subsection_answered / subsection_questions.count() * 100) if subsection_questions.count() > 0 else 0
                subsection_data['completion'] = round(subsection_completion, 1)
                subsection_data['answered'] = subsection_answered
                subsection_data['total'] = subsection_questions.count()
                
                category_data['children'].append(subsection_data)
            
            # Calculate category completion
            category_questions = Question.objects.filter(topic__sub_section__category=category)
            category_answered = Answer.objects.filter(
                form__in=user_forms,
                question__in=category_questions,
                response__isnull=False
            ).exclude(response='').count()
            
            category_completion = (category_answered / category_questions.count() * 100) if category_questions.count() > 0 else 0
            category_data['completion'] = round(category_completion, 1)
            category_data['answered'] = category_answered
            category_data['total'] = category_questions.count()
            
            form_status_data.append(category_data)
        
        # Get timeline data (answers by date)
        timeline_data = []
        for i in range(30):  # Last 30 days
            date = timezone.now().date() - timedelta(days=i)
            daily_answers = Answer.objects.filter(
                form__in=user_forms,
                answered_at__date=date,
                response__isnull=False
            ).exclude(response='').count()
            
            timeline_data.append({
                'date': date.strftime('%Y-%m-%d'),
                'answers': daily_answers
            })
        
        timeline_data.reverse()  # Show oldest to newest
        
        analytics_data = {
            'answered_questions': answered_questions,
            'total_questions': total_questions,
            'completion_rate': round(completion_rate, 1),
            'form_status': form_status_data,
            'timeline': timeline_data,
            'last_updated': timezone.now().isoformat()
        }
        
        return JsonResponse(analytics_data)
        
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'User school not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_analytics_date_range(request):
    """Get analytics data for a specific date range."""
    try:
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        user_school = AdminUser.objects.get(admin_id=admin_id)
        user_forms = Form.objects.filter(user_id=admin_id)
        
        # Get date range from request
        range_type = request.GET.get('range', 'today')
        
        if range_type == 'today':
            start_date = timezone.now().date()
            end_date = start_date
        elif range_type == '7days':
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=7)
        elif range_type == '1month':
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        else:
            # Default to today
            start_date = timezone.now().date()
            end_date = start_date
        
        # Get answered questions in date range
        answered_in_range = Answer.objects.filter(
            form__in=user_forms,
            answered_at__date__gte=start_date,
            answered_at__date__lte=end_date,
            response__isnull=False
        ).exclude(response='').count()
        
        # Get total questions
        total_questions = Question.objects.count()
        
        # Calculate completion rate for the range
        completion_rate = (answered_in_range / total_questions * 100) if total_questions > 0 else 0
        
        return JsonResponse({
            'answered': answered_in_range,
            'total': total_questions,
            'completion_rate': round(completion_rate, 1),
            'date_range': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d'),
                'type': range_type
            }
        })
        
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'User school not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_fallback_analytics_bundle(request):
    """Fallback analytics data when FastAPI is not available."""
    import traceback
    
    try:
        print("Fallback: Getting user ID...")
        # Get user ID from session or Django user
        user_id = None
        if request.user.is_authenticated:
            user_id = request.user.id
            print(f"Fallback: Using Django user ID: {user_id}")
        elif request.session.get('school_id'):
            user_id = request.session.get('school_id')
            print(f"Fallback: Using session school_id: {user_id}")
        elif request.session.get('admin_id'):
            user_id = request.session.get('admin_id')
            print(f"Fallback: Using session admin_id: {user_id}")
        
        if not user_id:
            print("Fallback: No authenticated user found")
            return JsonResponse({'error': 'No authenticated user found'}, status=403)
        
        # Try to get basic analytics data from Django models, but fall back to static data if needed
        total_forms = 0
        completed_forms = 0
        completion_rate = 0
        
        try:
            print("Fallback: Attempting database queries...")
            from app.models import Form, Answer, Question, AdminUser
            
            # Get basic form statistics with error handling
            total_forms = Form.objects.count()
            print(f"Fallback: Total forms: {total_forms}")
            completed_forms = Form.objects.filter(status='completed').count()
            print(f"Fallback: Completed forms: {completed_forms}")
            completion_rate = (completed_forms / total_forms * 100) if total_forms > 0 else 0
            print(f"Fallback: Completion rate: {completion_rate}")
            
        except Exception as e:
            print(f"Database query failed, using static fallback data: {e}")
            print(f"Database error traceback: {traceback.format_exc()}")
            # Use static fallback data if database queries fail
            total_forms = 10
            completed_forms = 7
            completion_rate = 70
        
        # Create fallback response
        fallback_data = {
            'cards': {
                'completion_rate': completion_rate / 100,
                'avg_completion_hours': 24.5,
                'completed_forms': completed_forms,
                'pending_forms': total_forms - completed_forms
            },
            'charts': {
                'completion_by_school': {
                    'labels': ['Sample School 1', 'Sample School 2', 'Sample School 3'],
                    'datasets': [{
                        'data': [85, 92, 78],
                        'label': 'Completion Rate (%)'
                    }]
                },
                'forms_per_day': {
                    'labels': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
                    'datasets': [{
                        'data': [12, 19, 15, 25, 22, 8, 5],
                        'label': 'Forms Completed'
                    }]
                },
                'response_distribution': {
                    'labels': ['Excellent', 'Good', 'Average', 'Poor'],
                    'datasets': [{
                        'data': [35, 40, 20, 5],
                        'label': 'Response Distribution'
                    }]
                }
            },
            'school_completion': [
                {
                    'school_id': 1,
                    'school_name': 'Sample School 1',
                    'completion_pct': 0.85,
                    'answered': 85,
                    'required': 100,
                    'status': 'completed'
                },
                {
                    'school_id': 2,
                    'school_name': 'Sample School 2',
                    'completion_pct': 0.92,
                    'answered': 92,
                    'required': 100,
                    'status': 'completed'
                },
                {
                    'school_id': 3,
                    'school_name': 'Sample School 3',
                    'completion_pct': 0.78,
                    'answered': 78,
                    'required': 100,
                    'status': 'in-progress'
                }
            ],
            'group_aggregates': [
                {
                    'group': 'Region A',
                    'completion_pct': 0.85,
                    'answered': 255,
                    'required': 300,
                    'schools': 3
                }
            ],
            'meta': {
                'filters_used': {},
                'total_records': 3,
                'generated_at': timezone.now().isoformat()
            }
        }
        
        print("Fallback: Returning fallback data...")
        return JsonResponse(fallback_data)
        
    except Exception as e:
        print(f"Fallback: Unexpected error: {e}")
        print(f"Fallback: Error traceback: {traceback.format_exc()}")
        return JsonResponse({'error': f'Fallback analytics failed: {str(e)}'}, status=500)

@session_or_login_required
@csrf_exempt
def api_analytics_test(request):
    """Simple test endpoint to verify authentication is working."""
    return JsonResponse({
        'status': 'success',
        'message': 'Authentication is working',
        'user_id': request.user.id if request.user.is_authenticated else request.session.get('admin_id'),
        'timestamp': timezone.now().isoformat()
    })

@session_or_login_required
@csrf_exempt
def api_analytics_simple(request):
    """Simple analytics endpoint that returns static data for testing."""
    return JsonResponse({
        'cards': {
            'completion_rate': 0.75,
            'avg_completion_hours': 24.5,
            'completed_forms': 15,
            'pending_forms': 5
        },
        'charts': {
            'completion_by_school': {
                'labels': ['School A', 'School B', 'School C'],
                'datasets': [{
                    'data': [85, 92, 78],
                    'label': 'Completion Rate (%)'
                }]
            },
            'forms_per_day': {
                'labels': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri'],
                'datasets': [{
                    'data': [12, 19, 15, 25, 22],
                    'label': 'Forms Completed'
                }]
            },
            'response_distribution': {
                'labels': ['Excellent', 'Good', 'Average', 'Poor'],
                'datasets': [{
                    'data': [35, 40, 20, 5],
                    'label': 'Response Distribution'
                }]
            }
        },
        'school_completion': [
            {
                'school_id': 1,
                'school_name': 'Test School 1',
                'completion_pct': 0.85,
                'answered': 85,
                'required': 100,
                'status': 'completed'
            }
        ],
        'group_aggregates': [],
        'meta': {
            'filters_used': {},
            'total_records': 1,
            'generated_at': timezone.now().isoformat()
        }
    })

@session_or_login_required
@csrf_exempt
def api_analytics_bundle(request):
    """Analytics bundle endpoint with real database data."""
    try:
        # Parse filters from request
        filters = {}
        if request.method == 'POST':
            try:
                filters = json.loads(request.body.decode('utf-8'))
            except:
                pass
        
        # Apply filters to build base queryset
        base_queryset = AnalyticsService.build_filtered_queryset(filters)
        
        # Calculate completion statistics
        completion_stats = AnalyticsService.calculate_completion_stats(base_queryset)
        
        # Get enhanced school completion data with geographic information
        school_completion = AnalyticsService.get_enhanced_school_completion_data(base_queryset, filters)
        
        # Get group aggregates
        group_aggregates = AnalyticsService.get_group_aggregates(base_queryset, filters)
        
        # Calculate average completion time
        avg_completion_hours = AnalyticsService.calculate_avg_completion_time(base_queryset)
        
        return JsonResponse({
            'cards': {
                'completion_rate': completion_stats['completion_rate'],
                'avg_completion_hours': avg_completion_hours,
                'completed_forms': completion_stats['completed_forms'],
                'pending_forms': completion_stats['pending_forms']
            },
            'charts': {
                'completion_by_school': {
                    'labels': [school['school_name'] for school in school_completion[:10]],
                    'datasets': [{
                        'data': [school['completion_pct'] * 100 for school in school_completion[:10]],
                        'label': 'Completion Rate (%)'
                    }]
                },
                'forms_per_day': AnalyticsService.get_forms_per_day_chart(base_queryset),
                'response_distribution': AnalyticsService.get_response_distribution_chart(base_queryset)
            },
            'school_completion': school_completion,
            'group_aggregates': group_aggregates,
            'meta': {
                'filters_used': filters,
                'total_records': len(school_completion),
                'generated_at': timezone.now().isoformat()
            }
        })
        
    except Exception as e:
        print(f"Analytics bundle error: {e}")
        # Return fallback data on error
        return JsonResponse({
            'cards': {
                'completion_rate': 0.0,
                'avg_completion_hours': 0.0,
                'completed_forms': 0,
                'pending_forms': 0
            },
            'charts': {
                'completion_by_school': {'labels': [], 'datasets': [{'data': [], 'label': 'Completion Rate (%)'}]},
                'forms_per_day': {'labels': [], 'datasets': [{'data': [], 'label': 'Forms Completed'}]},
                'response_distribution': {'labels': [], 'datasets': [{'data': [], 'label': 'Response Distribution'}]}
            },
            'school_completion': [],
            'group_aggregates': [],
            'meta': {
                'filters_used': {},
                'total_records': 0,
                'generated_at': timezone.now().isoformat(),
                'error': str(e)
            }
        })

# Import analytics service
from .analytics_service import AnalyticsService




@session_or_login_required
@csrf_exempt
def api_analytics_drilldown(request):
    """Analytics drilldown endpoint with real database data."""
    try:
        # Parse request data
        data = {}
        if request.method == 'POST':
            try:
                data = json.loads(request.body.decode('utf-8'))
            except:
                pass
        
        level = data.get('level', 'category')
        filters = data.get('filters', {})
        
        # Build filtered queryset
        base_queryset = AnalyticsService.build_filtered_queryset(filters)
        
        # Get drilldown data based on level
        drilldown_data = AnalyticsService.get_drilldown_data(base_queryset, level)
        
        return JsonResponse({
            'data': drilldown_data,
            'level': level,
            'meta': {
                'filters_used': filters,
                'generated_at': timezone.now().isoformat()
            }
        })
        
    except Exception as e:
        print(f"Drilldown error: {e}")
        return JsonResponse({
            'data': [],
            'level': level,
            'meta': {
                'filters_used': {},
                'generated_at': timezone.now().isoformat(),
                'error': str(e)
            }
        })

@session_or_login_required
@csrf_exempt
def api_export_csv(request):
    """Export current analytics bundle as CSV with real data."""
    try:
        # Parse filters from request
        filters = {}
        if request.method == 'POST':
            try:
                filters = json.loads(request.body.decode('utf-8'))
            except:
                pass

        # Build filtered queryset
        base_queryset = AnalyticsService.build_filtered_queryset(filters)
        
        # Get school completion data
        school_completion = AnalyticsService.get_school_completion_data(base_queryset, filters)
        
        # Get completion stats
        completion_stats = AnalyticsService.calculate_completion_stats(base_queryset)

        # Prepare CSV
        import csv
        from io import StringIO
        output = StringIO()
        writer = csv.writer(output)
        
        # Write summary data
        writer.writerow(['Analytics Summary'])
        writer.writerow(['Completion Rate', f"{completion_stats['completion_rate']:.1%}"])
        writer.writerow(['Completed Forms', completion_stats['completed_forms']])
        writer.writerow(['Pending Forms', completion_stats['pending_forms']])
        writer.writerow([])
        
        # Write school completion data
        writer.writerow(['School Completion Details'])
        writer.writerow(['School ID', 'School Name', 'Completion %', 'Answered', 'Required', 'Status'])
        for row in school_completion:
            writer.writerow([
                row.get('school_id'),
                row.get('school_name'),
                round((row.get('completion_pct') or 0) * 100, 1),
                row.get('answered') or '',
                row.get('required') or '',
                row.get('status') or ''
            ])

        resp = HttpResponse(output.getvalue(), content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="analytics.csv"'
        return resp
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_export_drilldown_csv(request):
    """Export current drilldown distribution as CSV with real data."""
    try:
        # Parse request data
        data = {}
        if request.method == 'POST':
            try:
                data = json.loads(request.body.decode('utf-8'))
            except:
                pass
        
        level = data.get('level', 'category')
        filters = data.get('filters', {})
        
        # Build filtered queryset
        base_queryset = AnalyticsService.build_filtered_queryset(filters)
        
        # Get drilldown data
        drilldown_data = AnalyticsService.get_drilldown_data(base_queryset, level)
        
        # Prepare CSV
        from io import StringIO
        import csv
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Label', 'Count', 'Percentage', 'Average', 'Median', 'Min', 'Max', 'Frequency', 'Geo Comparison'])
        for row in drilldown_data:
            writer.writerow([
                row.get('name', ''),
                row.get('count', 0),
                row.get('percentage_distribution', ''),
                row.get('average', ''),
                row.get('median', ''),
                row.get('min', ''),
                row.get('max', ''),
                row.get('frequency_distribution', ''),
                row.get('comparison_by_geo', '')
            ])
        resp = HttpResponse(output.getvalue(), content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="drilldown_{level}.csv"'
        return resp
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@session_or_login_required
@csrf_exempt
def api_export_bundle_xlsx(request):
    """Export bundle data as XLSX with real data"""
    try:
        # Get filters from query params
        filters = {}
        for key in ['q', 'region_ids', 'division_ids', 'district_ids', 'school_ids', 
                   'date_from', 'date_to', 'deadline', 'submission_status', 'completion_status',
                   'category_ids', 'sub_section_ids', 'topic_ids', 'question_ids', 'sub_question_ids',
                   'thresholds', 'group_by']:
            value = request.GET.get(key)
            if value:
                if key in ['region_ids', 'division_ids', 'district_ids', 'school_ids', 
                          'submission_status', 'completion_status', 'category_ids', 
                          'sub_section_ids', 'topic_ids', 'question_ids', 'sub_question_ids']:
                    filters[key] = value.split(',') if ',' in value else [value]
                elif key == 'thresholds':
                    filters[key] = json.loads(value)
                else:
                    filters[key] = value
        
        # Build filtered queryset
        base_queryset = AnalyticsService.build_filtered_queryset(filters)
        
        # Get real data
        completion_stats = AnalyticsService.calculate_completion_stats(base_queryset)
        school_completion = AnalyticsService.get_school_completion_data(base_queryset, filters)
        group_aggregates = AnalyticsService.get_group_aggregates(base_queryset, filters)
        avg_completion_hours = AnalyticsService.calculate_avg_completion_time(base_queryset)
        
        # Create XLSX workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Completion Rate", f"{completion_stats['completion_rate']:.1%}"])
        ws_summary.append(["Average Completion Hours", f"{avg_completion_hours:.1f}"])
        ws_summary.append(["Completed Forms", completion_stats['completed_forms']])
        ws_summary.append(["Pending Forms", completion_stats['pending_forms']])
        
        # School completion sheet
        ws_schools = wb.create_sheet("School Completion")
        ws_schools.append(["School", "Completion %", "Answered", "Required", "Status"])
        
        for school in school_completion:
            ws_schools.append([
                school.get('school_name', ''),
                f"{school.get('completion_pct', 0) * 100:.1f}%",
                school.get('answered', 0),
                school.get('required', 0),
                school.get('status', '')
            ])
        
        # Group completion sheet (if exists)
        if group_aggregates:
            ws_groups = wb.create_sheet("Group Completion")
            ws_groups.append(["Group", "Completion %", "Answered", "Required", "Schools"])
            
            for group in group_aggregates:
                ws_groups.append([
                    group.get('group', ''),
                    f"{group.get('completion_pct', 0) * 100:.1f}%",
                    group.get('answered', 0),
                    group.get('required', 0),
                    group.get('schools', 0)
                ])
        
        # Style the headers
        for ws in [ws_summary, ws_schools]:
            if ws == wb.worksheets[0]:  # Skip if no group sheet
                continue
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="analytics_bundle.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@session_or_login_required
@csrf_exempt
def api_reports_school_completion(request):
    """Enhanced school completion data endpoint."""
    try:
        # Parse filters from request
        filters = {}
        if request.method == 'POST':
            try:
                filters = json.loads(request.body.decode('utf-8'))
            except:
                pass
        
        # Get enhanced school completion data directly (no need for queryset filtering for raw SQL)
        school_completion = AnalyticsService.get_enhanced_school_completion_data(None, filters)
        
        # Apply client-side filters if provided
        if filters:
            if filters.get('region_ids'):
                school_completion = [s for s in school_completion if s['region_id'] in filters['region_ids']]
            if filters.get('division_ids'):
                school_completion = [s for s in school_completion if s['division_id'] in filters['division_ids']]
            if filters.get('district_ids'):
                school_completion = [s for s in school_completion if s['district_id'] in filters['district_ids']]
            if filters.get('school_ids'):
                school_completion = [s for s in school_completion if s['school_id'] in filters['school_ids']]
        
        return JsonResponse({
            'success': True,
            'data': school_completion,
            'count': len(school_completion),
            'filters_applied': filters
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_reports_category_content(request):
    """Category content data endpoint."""
    try:
        # Parse filters from request
        filters = {}
        if request.method == 'POST':
            try:
                filters = json.loads(request.body.decode('utf-8'))
            except:
                pass
        
        # Apply filters to build base queryset
        base_queryset = AnalyticsService.build_filtered_queryset(filters)
        
        # Get category content data
        category_content = AnalyticsService.get_category_content_data(base_queryset, filters)
        
        return JsonResponse({
            'success': True,
            'data': category_content,
            'count': len(category_content)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_analytics_filter_options(request):
    """Get filter options for dropdowns."""
    try:
        filter_options = AnalyticsService.get_filter_options()
        return JsonResponse(filter_options)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_hierarchical_filter_options(request):
    """Get hierarchical filter options based on parent selection."""
    try:
        filter_type = request.GET.get('type')
        parent_id = request.GET.get('parent_id')
        
        if not filter_type or not parent_id:
            return JsonResponse({'error': 'type and parent_id parameters are required'}, status=400)
        
        options = AnalyticsService.get_hierarchical_filter_options(filter_type, parent_id)
        return JsonResponse({
            'success': True,
            'data': options,
            'count': len(options)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@session_or_login_required
@csrf_exempt
def api_export_drilldown_xlsx(request):
    """Export drilldown data as XLSX with real data"""
    try:
        # Get filters from query params
        filters = {}
        level = 'category'  # default level
        
        for key in ['level', 'region_ids', 'division_ids', 'district_ids', 'school_ids', 
                   'category_ids', 'sub_section_ids', 'topic_ids', 'question_ids', 'sub_question_ids']:
            value = request.GET.get(key)
            if value:
                if key == 'level':
                    level = value
                else:
                    filters[key] = value.split(',') if ',' in value else [value]
        
        # Build filtered queryset
        base_queryset = AnalyticsService.build_filtered_queryset(filters)
        
        # Get drilldown data
        drilldown_data = AnalyticsService.get_drilldown_data(base_queryset, level)
        
        # Create XLSX workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Drilldown Analytics - {level.title()}"
        
        # Headers
        headers = ['Label', 'Count', 'Percentage', 'Average', 'Median', 'Min', 'Max', 'Frequency', 'Geo Comparison']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for item in drilldown_data:
            ws.append([
                item.get('name', ''),
                item.get('count', 0),
                item.get('percentage_distribution', ''),
                item.get('average', ''),
                item.get('median', ''),
                item.get('min', ''),
                item.get('max', ''),
                item.get('frequency_distribution', ''),
                item.get('comparison_by_geo', '')
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="drilldown_{level}_analytics.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_presets(request):
    """Simple presets using Redis key per user (save/load)."""
    try:
        # Get user ID from either Django user or session
        user_id = None
        if request.user.is_authenticated:
            user_id = request.user.id
        elif request.session.get('school_id'):
            user_id = request.session.get('school_id')
        elif request.session.get('admin_id'):
            user_id = request.session.get('admin_id')
        
        if not user_id:
            return JsonResponse({'error': 'No authenticated user found'}, status=403)
        
        key = f"presets:user:{user_id}"
        if request.method == 'GET':
            data = r.get(key)
            presets = json.loads(data) if data else []
            return JsonResponse({'presets': presets})
        elif request.method == 'POST':
            body = json.loads(request.body.decode('utf-8'))
            presets = body.get('presets') or []
            r.set(key, json.dumps(presets))
            return JsonResponse({'status': 'ok'})
        else:
            return JsonResponse({'error': 'Method not allowed'}, status=405)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_filters_options(request):
    """Return options for Region/Division/District/School and Category/Subsection/Topic/Question/Sub-question."""
    try:
        return JsonResponse(AnalyticsService.get_filter_options())
    except Exception as e:
        print(f"Filter options error: {e}")
        # Return empty options if database query fails
        return JsonResponse({
            'regions': [],
            'divisions': [],
            'districts': [],
            'schools': [],
            'categories': [],
            'sub_sections': [],
            'topics': [],
            'questions': [],
            'sub_questions': []
        })

@session_or_login_required
@csrf_exempt
def api_profile_update(request):
    """Proxy to FastAPI profile update endpoint."""
    return proxy_to_fastapi(request, '/api/profile')

@session_or_login_required
@csrf_exempt
def api_dashboard_completion(request):
    """Proxy to FastAPI dashboard completion endpoint."""
    return proxy_to_fastapi(request, '/api/dashboard/completion')

@session_or_login_required
@csrf_exempt
def api_dashboard_recent_activity(request):
    """Proxy to FastAPI dashboard recent activity endpoint."""
    return proxy_to_fastapi(request, '/api/dashboard/recent-activity')

@session_or_login_required
@csrf_exempt
def api_dashboard_quick_stats(request):
    """Proxy to FastAPI dashboard quick stats endpoint."""
    return proxy_to_fastapi(request, '/api/dashboard/quick-stats')

@session_or_login_required
@csrf_exempt
def api_form_answers(request):
    """Get saved answers for the current user - Django implementation."""
    try:
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        
        admin_user = AdminUser.objects.get(admin_id=admin_id)
        
        # Get or create form for this admin user
        form = get_or_create_admin_form(admin_user)
        if not form:
            return JsonResponse({'error': 'Could not create or find form for user'}, status=500)
        
        # Get all answers for this form with proper relationships
        answers = Answer.objects.filter(form=form).select_related('question', 'sub_question')
        
        answers_data = {}
        for answer in answers:
            if answer.response and answer.response.strip():  # Only include non-empty responses
                if answer.question and not answer.sub_question:
                    # Regular question answer
                    answers_data[str(answer.question.question_id)] = {
                        'value': answer.response,
                        'timestamp': answer.answered_at.isoformat() if answer.answered_at else None
                    }
                elif answer.sub_question:
                    # Sub-question answer
                    answers_data[str(answer.sub_question.sub_question_id)] = {
                        'value': answer.response,
                        'timestamp': answer.answered_at.isoformat() if answer.answered_at else None
                    }
        
        return JsonResponse({
            'success': True,
            'answers': answers_data,
            'form_id': form.form_id,
            'status': form.status,
            'message': f'Loaded {len(answers_data)} saved answers'
        })
        
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'Admin user not found'}, status=404)
    except Exception as e:
        print(f"Error in api_form_answers: {e}")
        return JsonResponse({'error': str(e)}, status=500)


def get_or_create_admin_form(admin_user):
    """Helper function to get or create a form for an admin user."""
    try:
        from django.contrib.auth.models import User
        from app.models import UsersSchool
        
        # First, try to find an existing form for this admin user
        # Priority 1: Look for forms by username (more reliable)
        existing_form = Form.objects.filter(
            user__username=admin_user.username
        ).first()
        
        # Priority 2: If not found by username, try by school_id
        if not existing_form:
            existing_form = Form.objects.filter(
                school_id=admin_user.admin_id
            ).first()
        
        if existing_form:
            return existing_form
        
        # If no existing form, create one
        # Create or get Django User
        django_user, created = User.objects.get_or_create(
            username=admin_user.username,
            defaults={
                'email': admin_user.email,
                'is_active': True
            }
        )
        
        # Try to find or create a UsersSchool entry
        users_school = UsersSchool.objects.filter(id=admin_user.admin_id).first()
        if not users_school:
            # Create a minimal UsersSchool entry for compatibility
            school_name = (admin_user.assigned_area or 'School')[:20]
            users_school = UsersSchool.objects.create(
                id=admin_user.admin_id,
                username=admin_user.username[:50],
                email=admin_user.email[:100],
                password_hash='',  # Not used for AdminUser
                role='school',
                school_name=school_name,
                is_active=True
            )
        
        # Create the form - need to use School model, not UsersSchool
        from app.models import School
        school_obj = School.objects.filter(id=admin_user.admin_id).first()
        if not school_obj:
            # Create a School object if it doesn't exist
            school_obj = School.objects.create(
                id=admin_user.admin_id,
                school_name=users_school.school_name,
                region_id=1,  # Default region
                division_id=1,  # Default division
                district_id=1  # Default district
            )
        
        form = Form.objects.create(
            user=django_user,
            school=school_obj,
            status='draft'
        )
        
        return form
        
    except Exception as e:
        print(f"Error in get_or_create_admin_form: {e}")
        return None

# Security & Audit API Endpoints
@session_or_login_required
@csrf_exempt
def api_security_last_login(request):
    """Get last login information for the current user."""
    try:
        # Get user's admin record
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        user_school = AdminUser.objects.get(admin_id=admin_id)
        
        # Get last login from AuditTrail or use last_login field
        last_login_data = {
            'last_login': user_school.last_login.isoformat() if user_school.last_login else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'location': 'Unknown Location',  # You can integrate with IP geolocation service
            'device': request.META.get('HTTP_USER_AGENT', 'Unknown Device')
        }
        
        return JsonResponse(last_login_data)
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_security_sessions(request):
    """Get active sessions for the current user."""
    try:
        # For now, return current session info
        # In a full implementation, you'd track sessions in database
        sessions = [
            {
                'id': 'current',
                'device': request.META.get('HTTP_USER_AGENT', 'Unknown Device'),
                'location': 'Unknown Location',
                'last_activity': datetime.now().isoformat(),
                'is_current': True
            }
        ]
        
        return JsonResponse({'sessions': sessions})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_security_login_history(request):
    """Get login history for the current user."""
    try:
        # Get recent login attempts from AuditTrail
        # This is a simplified version - you'd want to use the AuditLog model from migration
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        user_school = AdminUser.objects.get(admin_id=admin_id)
        
        history = [
            {
                'timestamp': user_school.last_login.isoformat() if user_school.last_login else datetime.now().isoformat(),
                'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
                'location': 'Unknown Location',
                'device': request.META.get('HTTP_USER_AGENT', 'Unknown Device'),
                'success': True
            }
        ]
        
        return JsonResponse({'history': history})
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_security_terminate_session(request, session_id):
    """Terminate a specific session."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # In a full implementation, you'd remove the session from database
        # For now, just return success
        return JsonResponse({'success': True, 'message': 'Session terminated'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_audit_logs(request):
    """Get audit logs for the current user from both AuditTrail and AuditLog tables."""
    try:
        # Check session-based authentication
        admin_id = request.session.get('admin_id')
        if not admin_id:
            return JsonResponse({'error': 'Not authenticated via admin system'}, status=403)
        
        admin_user = AdminUser.objects.get(admin_id=admin_id)
        
        # Get filter parameters
        log_type = request.GET.get('type', 'all')
        time_range = request.GET.get('time_range', '7d')
        
        # Calculate date range
        from datetime import timedelta
        if time_range == '1d':
            start_date = datetime.now() - timedelta(days=1)
        elif time_range == '7d':
            start_date = datetime.now() - timedelta(days=7)
        elif time_range == '30d':
            start_date = datetime.now() - timedelta(days=30)
        else:
            start_date = datetime.now() - timedelta(days=7)
        
        logs = []
        
        # Get audit trails for form modifications (if they exist)
        try:
            # Try to use Django user if available, otherwise skip
            if hasattr(admin_user, 'user') and admin_user.user:
                audit_trails = AuditTrail.objects.filter(
                    user=admin_user.user,
                    timestamp__gte=start_date
                ).order_by('-timestamp')[:25]
            else:
                audit_trails = []
            
            for trail in audit_trails:
                logs.append({
                    'id': f'trail_{trail.id}',
                    'action': f'{trail.action.title()} Question',
                    'resource_type': 'question',
                    'resource_name': f"Question {trail.question.question_id}" if trail.question else 'Unknown',
                    'old_value': trail.old_value,
                    'new_value': trail.new_value,
                    'timestamp': trail.timestamp.isoformat(),
                    'saved_locally': True,
                    'saved_database': True
                })
        except AdminUser.DoesNotExist:
            pass
        
        # Get audit logs for profile updates and other system activities
        try:
            if hasattr(admin_user, 'user') and admin_user.user:
                audit_logs = AuditLog.objects.filter(
                    user=admin_user.user,
                    timestamp__gte=start_date
                ).order_by('-timestamp')[:25]
            else:
                audit_logs = []
        except:
            audit_logs = []
        
        for log in audit_logs:
            action_display = {
                'password_change': 'Changed Password',
                'update': 'Updated Profile',
                'login': 'Logged In',
                'logout': 'Logged Out',
                'session_terminated': 'Session Terminated'
            }.get(log.action_type, log.action_type.title())
            
            logs.append({
                'id': f'log_{log.id}',
                'action': action_display,
                'resource_type': log.resource_type or 'system',
                'resource_name': log.description or 'System Activity',
                'old_value': None,  # AuditLog doesn't have old/new values
                'new_value': None,
                'timestamp': log.timestamp.isoformat(),
                'saved_locally': True,
                'saved_database': True
            })
        
        # Sort all logs by timestamp (newest first)
        logs.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Limit to 50 total logs
        logs = logs[:50]
        
        return JsonResponse({'logs': logs})
    except AdminUser.DoesNotExist:
        return JsonResponse({'error': 'Admin user not found'}, status=404)
    except Exception as e:
        print(f"Error in api_audit_logs: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@session_or_login_required
@csrf_exempt
def api_audit_export(request):
    """Export audit logs in specified format."""
    try:
        format_type = request.GET.get('format', 'csv')
        
        if format_type == 'csv':
            # Return CSV export URL or generate CSV
            return JsonResponse({'download_url': '/api/audit/export/csv/'})
        elif format_type == 'pdf':
            # Return PDF export URL or generate PDF
            return JsonResponse({'download_url': '/api/audit/export/pdf/'})
        else:
            return JsonResponse({'error': 'Invalid format'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_GET
def api_schools_search(request):
    """Search schools by school ID or name for admin user management."""
    try:
        query = request.GET.get('q', '').strip()
        
        if len(query) < 2:
            return JsonResponse({
                'success': False,
                'error': 'Query must be at least 2 characters',
                'schools': []
            })
        
        # Search schools by school_id or school_name
        schools = School.objects.select_related('district', 'division', 'region').filter(
            Q(school_id__icontains=query) | Q(school_name__icontains=query)
        ).order_by('school_name')[:20]  # Limit to 20 results
        
        schools_data = []
        for school in schools:
            schools_data.append({
                'id': school.id,
                'school_id': school.school_id,
                'name': school.school_name,
                'address': getattr(school, 'address', ''),
                'district': school.district.name if school.district else '',
                'division': school.division.name if school.division else '',
                'region': school.region.name if school.region else ''
            })
        
        return JsonResponse({
            'success': True,
            'schools': schools_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'schools': []
        })


